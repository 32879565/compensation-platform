from __future__ import annotations

from copy import deepcopy
from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.auth.bootstrap import seed_rbac
from app.auth.permissions import Perm
from app.core.security import hash_password
from app.models.audit import AuditLog
from app.models.auth import Permission, Role, RolePermission, User, UserOrgScope, UserRole
from app.models.employee import Department, Employee
from app.models.org import OrgType, OrgUnit
from app.models.payroll_batch import BatchStatus, PayrollBatch
from app.models.payroll_result import PayrollResult

pytestmark = pytest.mark.usefixtures("pg_engine")


@pytest.fixture
def client(db_session):
    from fastapi.testclient import TestClient

    from app.db.session import get_session
    from app.main import app

    def _override():
        yield db_session

    app.dependency_overrides[get_session] = _override
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.clear()


def _user(session, username: str, roles: list[str], scope_ids: list[int] | None = None) -> User:
    seed_rbac(session)
    user = User(username=username, password_hash=hash_password("StrongPass123!"))
    session.add(user)
    session.flush()
    for role_code in roles:
        role = session.scalars(select(Role).where(Role.code == role_code)).one()
        session.add(UserRole(user_id=user.id, role_id=role.id))
    for org_unit_id in scope_ids or []:
        session.add(UserOrgScope(user_id=user.id, org_unit_id=org_unit_id))
    session.flush()
    return user


def _token(client, username: str) -> dict[str, str]:
    response = client.post(
        "/api/auth/login", json={"username": username, "password": "StrongPass123!"}
    )
    assert response.status_code == 200
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def _scoped_exporter_role(session) -> None:
    seed_rbac(session)
    role = Role(code="SCOPED_EXPORTER", name="Scoped exporter", is_global_scope=False)
    session.add(role)
    session.flush()
    permission_id = session.scalars(
        select(Permission.id).where(Permission.code == Perm.EXPORT_DATA)
    ).one()
    session.add(RolePermission(role_id=role.id, permission_id=permission_id))
    session.flush()


def _result(
    batch: PayrollBatch,
    employee: Employee,
    gross: str,
    *,
    version: int = 1,
    lines: list[dict] | None = None,
    input_snapshot: dict | None = None,
) -> PayrollResult:
    return PayrollResult(
        batch_id=batch.id,
        batch_version=batch.version,
        employee_id=employee.id,
        version=version,
        org_unit_id=employee.org_unit_id,
        department=employee.department,
        emp_no_snapshot=employee.emp_no,
        employee_name_snapshot=employee.name,
        id_card_snapshot=employee.id_card,
        bank_account_snapshot=employee.bank_account,
        social_city_snapshot=employee.social_city,
        actual_attendance_days=Decimal("22"),
        gross=Decimal(gross),
        deposit=Decimal("0"),
        net=Decimal(gross),
        carry_forward=Decimal("0"),
        rule_version="v2",
        input_snapshot=input_snapshot or {},
        lines=lines or [],
        exceptions=[],
        warnings=[],
        has_error=False,
    )


def _seed(session):
    group = OrgUnit(code="EXPORT_GROUP", name="Group", type=OrgType.GROUP)
    session.add(group)
    session.flush()
    visible_store = OrgUnit(
        code="@VISIBLE", name="=Visible Store", type=OrgType.STORE, parent_id=group.id
    )
    hidden_store = OrgUnit(
        code="HIDDEN", name="Hidden Store", type=OrgType.STORE, parent_id=group.id
    )
    session.add_all([visible_store, hidden_store])
    session.flush()
    visible_employee = Employee(
        emp_no="+VISIBLE-1",
        name="\x01@Visible employee",
        org_unit_id=visible_store.id,
        department=Department.DINING,
        social_city="Guangzhou",
        id_card="=VISIBLE-ID",
        bank_account="@VISIBLE-BANK",
    )
    hidden_employee = Employee(
        emp_no="HIDDEN-1",
        name="Hidden employee",
        org_unit_id=hidden_store.id,
        department=Department.KITCHEN,
    )
    session.add_all([visible_employee, hidden_employee])
    batch = PayrollBatch(
        period="2026-07",
        attendance_start=date(2026, 6, 26),
        attendance_end=date(2026, 7, 25),
        status=BatchStatus.LOCKED,
        version=1,
    )
    session.add(batch)
    session.flush()
    session.add_all(
        [
            _result(batch, visible_employee, "100", version=1),
            _result(
                batch,
                visible_employee,
                "120",
                version=2,
                lines=[
                    {"code": "SOCIAL_PENSION_EMPLOYEE", "amount": "-120"},
                    {"code": "SOCIAL_PENSION_EMPLOYER", "amount": "240"},
                    {"code": "HOUSING_FUND_EMPLOYEE", "amount": "-50"},
                    {"code": "HOUSING_FUND_EMPLOYER", "amount": "100"},
                    {"code": "IIT_WITHHOLDING", "amount": "-80"},
                ],
                input_snapshot={
                    "tax_withholding": {
                        "current_taxable_income": "5000",
                        "current_employee_contribution": "170",
                        "current_tax_withheld": "80",
                    },
                    "social_contributions": {
                        "PENSION": {"employee": "120", "employer": "240"},
                        "MEDICAL": {"employee": "0", "employer": "0"},
                        "UNEMPLOYMENT": {"employee": "0", "employer": "0"},
                        "WORK_INJURY": {"employee": "0", "employer": "0"},
                        "MATERNITY": {"employee": "0", "employer": "0"},
                        "HOUSING": {"employee": "50", "employer": "100"},
                    },
                },
            ),
            _result(batch, hidden_employee, "200"),
        ]
    )
    session.commit()
    return {"visible_store": visible_store, "visible_employee": visible_employee}


def _scoped_pii_exporter_role(session) -> None:
    seed_rbac(session)
    role = Role(code="SCOPED_PII_EXPORTER", name="Scoped PII exporter", is_global_scope=False)
    session.add(role)
    session.flush()
    permission_ids = list(
        session.scalars(
            select(Permission.id).where(Permission.code.in_([Perm.EXPORT_DATA, Perm.EMPLOYEE_PII]))
        ).all()
    )
    session.add_all(
        [
            RolePermission(role_id=role.id, permission_id=permission_id)
            for permission_id in permission_ids
        ]
    )
    session.flush()


def test_scoped_payroll_export_is_formula_safe_and_audited(client, db_session):
    orgs = _seed(db_session)
    _scoped_exporter_role(db_session)
    exporter = _user(db_session, "exporter", ["SCOPED_EXPORTER"], [orgs["visible_store"].id])

    response = client.get(
        "/api/exports/payroll",
        headers=_token(client, exporter.username),
        params={"period": "2026-07"},
    )
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["cache-control"] == "no-store"
    sheet = load_workbook(BytesIO(response.content), data_only=False).active
    assert sheet.max_row == 2
    assert sheet["B2"].value == "'+VISIBLE-1"
    assert sheet["C2"].value == "'@Visible employee"
    assert sheet["D2"].value == "'@VISIBLE"
    assert sheet["E2"].value == "'=Visible Store"
    assert sheet["H2"].value == 120
    audit_row = db_session.scalars(
        select(AuditLog).where(
            AuditLog.actor_user_id == exporter.id, AuditLog.action == "export.payroll"
        )
    ).one()
    assert audit_row.detail == {"period": "2026-07", "rows": 1, "format": "xlsx"}


def test_unrelated_global_role_does_not_widen_scoped_export(client, db_session):
    orgs = _seed(db_session)
    _scoped_exporter_role(db_session)
    # AUDITOR is global but intentionally has no export:data.  The export
    # grant comes only from the scoped role and must remain store-limited.
    exporter = _user(
        db_session,
        "mixed-exporter",
        ["AUDITOR", "SCOPED_EXPORTER"],
        [orgs["visible_store"].id],
    )

    response = client.get(
        "/api/exports/payroll",
        headers=_token(client, exporter.username),
        params={"period": "2026-07"},
    )
    assert response.status_code == 200, response.text
    sheet = load_workbook(BytesIO(response.content), data_only=False).active
    assert sheet.max_row == 2
    assert sheet["B2"].value == "'+VISIBLE-1"


def test_payroll_export_requires_export_permission(client, db_session):
    _seed(db_session)
    _user(db_session, "ordinary", ["EMPLOYEE"])

    assert (
        client.get(
            "/api/exports/payroll", headers=_token(client, "ordinary"), params={"period": "2026-07"}
        ).status_code
        == 403
    )


@pytest.mark.parametrize(
    "endpoint",
    ["social-insurance", "individual-income-tax", "bank-payment"],
)
def test_regulatory_exports_require_pii_permission(client, db_session, endpoint):
    _seed(db_session)
    _user(db_session, "finance", ["FINANCE"])

    response = client.get(
        f"/api/exports/{endpoint}",
        headers=_token(client, "finance"),
        params={"period": "2026-07"},
    )

    assert response.status_code == 403


def test_scoped_regulatory_exports_are_pii_safe_formula_safe_and_audited(client, db_session):
    orgs = _seed(db_session)
    _scoped_pii_exporter_role(db_session)
    exporter = _user(
        db_session,
        "pii-exporter",
        ["SCOPED_PII_EXPORTER"],
        [orgs["visible_store"].id],
    )
    headers = _token(client, exporter.username)

    social = client.get(
        "/api/exports/social-insurance", headers=headers, params={"period": "2026-07"}
    )
    assert social.status_code == 200, social.text
    assert social.headers["cache-control"] == "no-store"
    social_sheet = load_workbook(BytesIO(social.content), data_only=False).active
    assert social_sheet.max_row == 2
    assert social_sheet["B2"].value == "'+VISIBLE-1"
    assert social_sheet["C2"].value == "'@Visible employee"
    assert social_sheet["D2"].value == "'=VISIBLE-ID"
    assert social_sheet["F2"].value == 120
    assert social_sheet["G2"].value == 240
    assert social_sheet["H2"].value == 50
    assert social_sheet["I2"].value == 100

    tax = client.get(
        "/api/exports/individual-income-tax", headers=headers, params={"period": "2026-07"}
    )
    assert tax.status_code == 200, tax.text
    assert tax.headers["cache-control"] == "no-store"
    tax_sheet = load_workbook(BytesIO(tax.content), data_only=False).active
    assert tax_sheet.max_row == 2
    assert tax_sheet["F2"].value == 5000
    assert tax_sheet["G2"].value == 170
    assert tax_sheet["H2"].value == 80

    bank = client.get("/api/exports/bank-payment", headers=headers, params={"period": "2026-07"})
    assert bank.status_code == 200, bank.text
    assert bank.headers["cache-control"] == "no-store"
    bank_sheet = load_workbook(BytesIO(bank.content), data_only=False).active
    assert bank_sheet.max_row == 2
    assert bank_sheet["D2"].value == "'@VISIBLE-BANK"
    assert bank_sheet["E2"].value == 120

    audit_rows = list(
        db_session.scalars(
            select(AuditLog)
            .where(AuditLog.actor_user_id == exporter.id, AuditLog.action.like("export.%"))
            .order_by(AuditLog.id)
        ).all()
    )
    assert [row.action for row in audit_rows] == [
        "export.social_insurance",
        "export.individual_income_tax",
        "export.bank_payment",
    ]
    assert all(
        row.detail == {"period": "2026-07", "rows": 1, "format": "generic-xlsx"}
        for row in audit_rows
    )
    assert all("VISIBLE" not in str(row.detail) for row in audit_rows)


def test_bank_payment_uses_locked_account_after_employee_bank_change(client, db_session):
    orgs = _seed(db_session)
    _scoped_pii_exporter_role(db_session)
    exporter = _user(
        db_session,
        "immutable-bank-exporter",
        ["SCOPED_PII_EXPORTER"],
        [orgs["visible_store"].id],
    )
    orgs["visible_employee"].bank_account = "ATTACKER-ACCOUNT"
    db_session.commit()

    response = client.get(
        "/api/exports/bank-payment",
        headers=_token(client, exporter.username),
        params={"period": "2026-07"},
    )

    assert response.status_code == 200, response.text
    sheet = load_workbook(BytesIO(response.content), data_only=False).active
    assert sheet["D2"].value == "'@VISIBLE-BANK"
    assert "ATTACKER" not in str(sheet["D2"].value)


def test_bank_payment_export_fails_closed_when_an_authorized_employee_lacks_an_account(
    client, db_session
):
    orgs = _seed(db_session)
    _scoped_pii_exporter_role(db_session)
    result = db_session.scalars(
        select(PayrollResult)
        .where(PayrollResult.employee_id == orgs["visible_employee"].id)
        .order_by(PayrollResult.version.desc())
    ).first()
    assert result is not None
    result.bank_account_snapshot = None
    db_session.commit()
    exporter = _user(
        db_session,
        "pii-exporter",
        ["SCOPED_PII_EXPORTER"],
        [orgs["visible_store"].id],
    )

    response = client.get(
        "/api/exports/bank-payment",
        headers=_token(client, exporter.username),
        params={"period": "2026-07"},
    )

    assert response.status_code == 422
    assert "bank account" in response.json()["detail"].lower()


@pytest.mark.parametrize(
    ("endpoint", "mutate_snapshot", "detail"),
    [
        (
            "social-insurance",
            lambda snapshot: snapshot.pop("social_contributions"),
            "social contribution snapshot",
        ),
        (
            "individual-income-tax",
            lambda snapshot: snapshot["tax_withholding"].pop("current_tax_withheld"),
            "tax value",
        ),
    ],
)
def test_regulatory_exports_fail_closed_for_missing_structured_locked_result_data(
    client, db_session, endpoint, mutate_snapshot, detail
):
    orgs = _seed(db_session)
    _scoped_pii_exporter_role(db_session)
    result = db_session.scalars(
        select(PayrollResult)
        .where(PayrollResult.employee_id == orgs["visible_employee"].id)
        .order_by(PayrollResult.version.desc())
    ).first()
    assert result is not None
    snapshot = deepcopy(result.input_snapshot)
    mutate_snapshot(snapshot)
    result.input_snapshot = snapshot
    db_session.commit()
    exporter = _user(
        db_session,
        "pii-exporter",
        ["SCOPED_PII_EXPORTER"],
        [orgs["visible_store"].id],
    )

    response = client.get(
        f"/api/exports/{endpoint}",
        headers=_token(client, exporter.username),
        params={"period": "2026-07"},
    )

    assert response.status_code == 409
    assert detail in response.json()["detail"].lower()
