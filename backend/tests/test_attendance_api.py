import io
from datetime import date
from decimal import Decimal

import pytest
from openpyxl import Workbook
from sqlalchemy import select

from app.auth.bootstrap import seed_rbac
from app.core.security import hash_password
from app.models.attendance import AttendanceRecord, ExpectedAttendanceRule, PerformanceRecord
from app.models.audit import AuditLog
from app.models.auth import Permission, Role, RolePermission, User, UserOrgScope, UserRole
from app.models.employee import Department, Employee
from app.models.org import OrgType, OrgUnit
from app.models.payroll_batch import BatchStatus, PayrollBatch
from app.models.payroll_result import PayrollResult

pytestmark = pytest.mark.usefixtures("pg_engine")


def _orgs(session):
    group = OrgUnit(code="G", name="集团", type=OrgType.GROUP)
    session.add(group)
    session.flush()
    gz = OrgUnit(code="R_GZ", name="广州", type=OrgType.REGION, parent_id=group.id)
    sz = OrgUnit(code="R_SZ", name="深圳", type=OrgType.REGION, parent_id=group.id)
    session.add_all([gz, sz])
    session.flush()
    gzs = OrgUnit(code="S_GZ", name="广州店", type=OrgType.STORE, parent_id=gz.id, city="广州")
    szs = OrgUnit(code="S_SZ", name="深圳店", type=OrgType.STORE, parent_id=sz.id, city="深圳")
    session.add_all([gzs, szs])
    session.flush()
    return gz, gzs, szs


def _emp(session, emp_no, org_id):
    e = Employee(emp_no=emp_no, name=emp_no, org_unit_id=org_id)
    session.add(e)
    session.flush()
    return e


def _global_expected_days_rule(session, *, monthly_expected_days: str = "22"):
    rule = ExpectedAttendanceRule(
        name="Global monthly expected-days rule",
        weekly_rest_days=[],
        monthly_expected_days=Decimal(monthly_expected_days),
        effective_from=date(2026, 5, 1),
        is_active=True,
    )
    session.add(rule)
    session.flush()
    return rule


def _user(session, username, roles, scope_ids=()):
    seed_rbac(session)
    u = User(username=username, password_hash=hash_password("StrongPass123!"))
    session.add(u)
    session.flush()
    for code in roles:
        role = session.scalars(select(Role).where(Role.code == code)).one()
        session.add(UserRole(user_id=u.id, role_id=role.id))
    for oid in scope_ids:
        session.add(UserOrgScope(user_id=u.id, org_unit_id=oid))
    session.flush()
    return u


def _scoped_schedule_role(session):
    from app.auth.permissions import Perm

    seed_rbac(session)
    role = Role(code="SCOPED_SCHEDULE", name="Scoped schedule", is_global_scope=False)
    session.add(role)
    session.flush()
    for code in (Perm.ATTENDANCE_SCHEDULE_READ, Perm.ATTENDANCE_SCHEDULE_WRITE):
        permission_id = session.scalars(select(Permission.id).where(Permission.code == code)).one()
        session.add(RolePermission(role_id=role.id, permission_id=permission_id))
    session.flush()


def _mixed_attendance_correction_roles(session):
    from app.auth.permissions import Perm

    seed_rbac(session)
    global_writer = Role(
        code="GLOBAL_ATTENDANCE_ONLY",
        name="Global attendance only",
        is_global_scope=True,
    )
    scoped_corrector = Role(
        code="SCOPED_PAYROLL_CORRECT_ONLY",
        name="Scoped payroll correct only",
        is_global_scope=False,
    )
    session.add_all([global_writer, scoped_corrector])
    session.flush()
    permission_ids = dict(
        session.execute(
            select(Permission.code, Permission.id).where(
                Permission.code.in_((Perm.ATTENDANCE_WRITE, Perm.PAYROLL_CORRECT))
            )
        ).all()
    )
    session.add_all(
        [
            RolePermission(
                role_id=global_writer.id,
                permission_id=permission_ids[Perm.ATTENDANCE_WRITE],
            ),
            RolePermission(
                role_id=scoped_corrector.id,
                permission_id=permission_ids[Perm.PAYROLL_CORRECT],
            ),
        ]
    )
    session.flush()


@pytest.fixture
def client(db_session):
    from fastapi.testclient import TestClient

    from app.db.session import get_session
    from app.main import app

    def _override():
        yield db_session

    app.dependency_overrides[get_session] = _override
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.clear()


def _token(client, username):
    r = client.post("/api/auth/login", json={"username": username, "password": "StrongPass123!"})
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


def test_set_and_list_attendance(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    emp = _emp(db_session, "E1", gzs.id)
    _global_expected_days_rule(db_session)
    _user(db_session, "hr", ["GROUP_HR"])
    h = _token(client, "hr")
    r = client.put(
        f"/api/employees/{emp.id}/attendance/2026-05",
        headers=h,
        json={"expected_days": "22", "actual_days": "21.5", "overtime_hours": "8"},
    )
    assert r.status_code == 200
    assert r.json()["actual_days"] == "21.50"
    updated = client.put(
        f"/api/employees/{emp.id}/attendance/2026-05",
        headers=h,
        json={"expected_days": "22", "actual_days": "20", "overtime_hours": "8"},
    )
    assert updated.status_code == 200, updated.text
    audit_row = db_session.scalars(
        select(AuditLog)
        .where(AuditLog.action == "attendance.set", AuditLog.target_id == emp.id)
        .order_by(AuditLog.id.desc())
    ).first()
    assert audit_row is not None
    assert audit_row.detail["before"]["actual_days"] == "21.50"
    assert audit_row.detail["after"]["actual_days"] == "20.00"
    lst = client.get("/api/attendance?period=2026-05", headers=h)
    assert lst.status_code == 200 and len(lst.json()) == 1


def test_scoped_schedule_writer_cannot_manage_global_or_other_org_rules(client, db_session):
    _gz, gz_store, sz_store = _orgs(db_session)
    _scoped_schedule_role(db_session)
    _user(db_session, "hr", ["GROUP_HR"])
    _user(db_session, "schedule-gz", ["SCOPED_SCHEDULE"], scope_ids=[gz_store.id])
    hr_headers = _token(client, "hr")
    scoped_headers = _token(client, "schedule-gz")

    def create_rule(name, org_unit_id):
        response = client.post(
            "/api/attendance-schedules",
            headers=hr_headers,
            json={
                "name": name,
                "org_unit_id": org_unit_id,
                "weekly_rest_days": [5, 6],
                "effective_from": "2026-01-01",
            },
        )
        assert response.status_code == 201, response.text
        return response.json()

    global_rule = create_rule("Global", None)
    gz_rule = create_rule("Guangzhou", gz_store.id)
    create_rule("Shenzhen", sz_store.id)

    listed = client.get("/api/attendance-schedules", headers=scoped_headers)
    assert listed.status_code == 200
    assert {rule["name"] for rule in listed.json()} == {"Global", "Guangzhou"}

    global_create = client.post(
        "/api/attendance-schedules",
        headers=scoped_headers,
        json={
            "name": "Scoped global bypass",
            "org_unit_id": None,
            "monthly_expected_days": "22",
            "effective_from": "2026-01-01",
        },
    )
    assert global_create.status_code == 403

    global_update = client.put(
        f"/api/attendance-schedules/{global_rule['id']}",
        headers=scoped_headers,
        json={**global_rule, "name": "Tampered global"},
    )
    assert global_update.status_code == 404

    own_update = client.put(
        f"/api/attendance-schedules/{gz_rule['id']}",
        headers=scoped_headers,
        json={**gz_rule, "name": "Scoped own rule"},
    )
    assert own_update.status_code == 200


def test_attendance_upsert_overwrites(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    emp = _emp(db_session, "E1", gzs.id)
    _global_expected_days_rule(db_session)
    _user(db_session, "hr", ["GROUP_HR"])
    h = _token(client, "hr")
    client.put(
        f"/api/employees/{emp.id}/attendance/2026-05",
        headers=h,
        json={"expected_days": "22", "actual_days": "20"},
    )
    client.put(
        f"/api/employees/{emp.id}/attendance/2026-05",
        headers=h,
        json={"expected_days": "22", "actual_days": "22"},
    )
    lst = client.get("/api/attendance?period=2026-05", headers=h).json()
    assert len(lst) == 1  # 未新增，覆盖
    assert lst[0]["actual_days"] == "22.00"


def test_new_attendance_rejects_missing_expected_days_rule(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    employee = _emp(db_session, "NO-RULE", gzs.id)
    _user(db_session, "hr", ["GROUP_HR"])
    headers = _token(client, "hr")

    response = client.put(
        f"/api/employees/{employee.id}/attendance/2026-05",
        headers=headers,
        json={"expected_days": "22", "actual_days": "20"},
    )

    assert response.status_code == 422
    assert (
        db_session.scalars(
            select(AttendanceRecord).where(AttendanceRecord.employee_id == employee.id)
        ).one_or_none()
        is None
    )


def test_expected_days_adjustment_requires_a_new_reason_and_is_audited(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    emp = _emp(db_session, "E1", gzs.id)
    _global_expected_days_rule(db_session)
    _user(db_session, "hr", ["GROUP_HR"])
    h = _token(client, "hr")
    base = {"actual_days": "22", "worked_hours": "0"}

    assert (
        client.put(
            f"/api/employees/{emp.id}/attendance/2026-05",
            headers=h,
            json={"expected_days": "22", **base},
        ).status_code
        == 200
    )
    assert (
        client.put(
            f"/api/employees/{emp.id}/attendance/2026-05",
            headers=h,
            json={
                "expected_days": "21",
                "expected_days_adjust_reason": "Roster correction",
                **base,
            },
        ).status_code
        == 200
    )
    assert (
        client.put(
            f"/api/employees/{emp.id}/attendance/2026-05",
            headers=h,
            json={
                "expected_days": "20",
                "expected_days_adjust_reason": "Roster correction",
                **base,
            },
        ).status_code
        == 422
    )
    assert (
        client.put(
            f"/api/employees/{emp.id}/attendance/2026-05",
            headers=h,
            json={
                "expected_days": "20",
                "expected_days_adjust_reason": "Second roster correction",
                **base,
            },
        ).status_code
        == 200
    )
    adjustment = db_session.scalars(
        select(AuditLog)
        .where(
            AuditLog.action == "attendance.expected_days.adjust",
            AuditLog.target_id == emp.id,
        )
        .order_by(AuditLog.id.desc())
    ).first()
    assert adjustment is not None
    assert adjustment.detail["reason"] == "Second roster correction"
    assert adjustment.detail["before"]["expected_days"] == "21.00"
    assert adjustment.detail["after"]["expected_days"] == "20.00"


def test_store_manager_cannot_override_a_generated_expected_days_value_even_with_reason(
    client, db_session
):
    _gz, gzs, _szs = _orgs(db_session)
    employee = _emp(db_session, "E1", gzs.id)
    _global_expected_days_rule(db_session)
    _user(db_session, "mgr", ["STORE_MANAGER"], scope_ids=[gzs.id])
    headers = _token(client, "mgr")

    response = client.put(
        f"/api/employees/{employee.id}/attendance/2026-05",
        headers=headers,
        json={
            "expected_days": "20",
            "expected_days_adjust_reason": "Manager roster exception",
            "actual_days": "20",
        },
    )

    assert response.status_code == 403
    record = db_session.scalars(
        select(AttendanceRecord).where(
            AttendanceRecord.employee_id == employee.id,
            AttendanceRecord.period == "2026-05",
        )
    ).one_or_none()
    assert record is None


def test_hr_adjustment_preserves_generated_expected_days_rule_and_audit(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    employee = _emp(db_session, "E1", gzs.id)
    rule = _global_expected_days_rule(db_session)
    _user(db_session, "hr", ["GROUP_HR"])
    headers = _token(client, "hr")

    response = client.put(
        f"/api/employees/{employee.id}/attendance/2026-05",
        headers=headers,
        json={
            "expected_days": "20",
            "expected_days_adjust_reason": "HR-approved roster exception",
            "actual_days": "20",
        },
    )

    assert response.status_code == 200, response.text
    assert response.json()["expected_days"] == "20.00"
    record = db_session.scalars(
        select(AttendanceRecord).where(
            AttendanceRecord.employee_id == employee.id,
            AttendanceRecord.period == "2026-05",
        )
    ).one()
    assert record.expected_days == Decimal("20")
    assert record.generated_expected_days == Decimal("22")
    assert record.expected_days_rule_id == rule.id
    assert record.expected_days_adjust_reason == "HR-approved roster exception"
    adjustment = db_session.scalars(
        select(AuditLog)
        .where(
            AuditLog.action == "attendance.expected_days.adjust",
            AuditLog.target_id == employee.id,
        )
        .order_by(AuditLog.id.desc())
    ).one()
    assert adjustment.detail["reason"] == "HR-approved roster exception"
    assert adjustment.detail["after"]["expected_days"] == "20.00"


def test_normal_attendance_entry_uses_generated_expected_days_over_client_value(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    employee = _emp(db_session, "E1", gzs.id)
    rule = _global_expected_days_rule(db_session)
    _user(db_session, "mgr", ["STORE_MANAGER"], scope_ids=[gzs.id])
    headers = _token(client, "mgr")

    response = client.put(
        f"/api/employees/{employee.id}/attendance/2026-05",
        headers=headers,
        json={"expected_days": "20", "actual_days": "20"},
    )

    assert response.status_code == 200, response.text
    assert response.json()["expected_days"] == "22.00"
    assert response.json()["expected_days_adjust_reason"] is None
    record = db_session.scalars(
        select(AttendanceRecord).where(
            AttendanceRecord.employee_id == employee.id,
            AttendanceRecord.period == "2026-05",
        )
    ).one()
    assert record.expected_days == Decimal("22")
    assert record.generated_expected_days == Decimal("22")
    assert record.expected_days_rule_id == rule.id
    assert record.expected_days_adjust_reason is None


def test_updating_attendance_keeps_the_persisted_schedule_basis(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    employee = _emp(db_session, "E1", gzs.id)
    rule = _global_expected_days_rule(db_session)
    _user(db_session, "mgr", ["STORE_MANAGER"], scope_ids=[gzs.id])
    headers = _token(client, "mgr")

    created = client.put(
        f"/api/employees/{employee.id}/attendance/2026-05",
        headers=headers,
        json={"expected_days": "20", "actual_days": "20"},
    )
    assert created.status_code == 200, created.text
    rule.monthly_expected_days = Decimal("21")
    db_session.flush()

    updated = client.put(
        f"/api/employees/{employee.id}/attendance/2026-05",
        headers=headers,
        json={"expected_days": "19", "actual_days": "19"},
    )

    assert updated.status_code == 200, updated.text
    assert updated.json()["expected_days"] == "22.00"
    record = db_session.scalars(
        select(AttendanceRecord).where(
            AttendanceRecord.employee_id == employee.id,
            AttendanceRecord.period == "2026-05",
        )
    ).one()
    assert record.expected_days == Decimal("22")
    assert record.generated_expected_days == Decimal("22")
    assert record.expected_days_rule_id == rule.id


def test_updating_legacy_attendance_does_not_backfill_schedule_provenance(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    employee = _emp(db_session, "E1", gzs.id)
    _global_expected_days_rule(db_session)
    db_session.add(
        AttendanceRecord(
            employee_id=employee.id,
            period="2026-05",
            expected_days=Decimal("20"),
            actual_days=Decimal("19"),
        )
    )
    db_session.flush()
    _user(db_session, "mgr", ["STORE_MANAGER"], scope_ids=[gzs.id])
    headers = _token(client, "mgr")

    updated = client.put(
        f"/api/employees/{employee.id}/attendance/2026-05",
        headers=headers,
        json={"expected_days": "22", "actual_days": "18"},
    )

    assert updated.status_code == 200, updated.text
    assert updated.json()["expected_days"] == "20.00"
    assert updated.json()["generated_expected_days"] is None
    record = db_session.scalars(
        select(AttendanceRecord).where(
            AttendanceRecord.employee_id == employee.id,
            AttendanceRecord.period == "2026-05",
        )
    ).one()
    assert record.expected_days == Decimal("20")
    assert record.generated_expected_days is None
    assert record.expected_days_rule_id is None


def test_hr_must_generate_schedule_before_adjusting_legacy_expected_days(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    employee = _emp(db_session, "E1", gzs.id)
    _global_expected_days_rule(db_session)
    db_session.add(
        AttendanceRecord(
            employee_id=employee.id,
            period="2026-05",
            expected_days=Decimal("20"),
            actual_days=Decimal("19"),
        )
    )
    db_session.flush()
    _user(db_session, "hr", ["GROUP_HR"])
    headers = _token(client, "hr")

    response = client.put(
        f"/api/employees/{employee.id}/attendance/2026-05",
        headers=headers,
        json={
            "expected_days": "21",
            "actual_days": "19",
            "expected_days_adjust_reason": "Review legacy roster basis",
        },
    )

    assert response.status_code == 422
    record = db_session.scalars(
        select(AttendanceRecord).where(
            AttendanceRecord.employee_id == employee.id,
            AttendanceRecord.period == "2026-05",
        )
    ).one()
    assert record.expected_days == Decimal("20")
    assert record.generated_expected_days is None
    assert record.expected_days_adjust_reason is None


def test_schedule_generation_resets_untrusted_legacy_expected_days_reason(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    employee = _emp(db_session, "E1", gzs.id)
    employee.hire_date = date(2026, 5, 1)
    rule = _global_expected_days_rule(db_session)
    db_session.add(
        AttendanceRecord(
            employee_id=employee.id,
            period="2026-05",
            expected_days=Decimal("20"),
            expected_days_adjust_reason="legacy free-text reason",
            actual_days=Decimal("19"),
        )
    )
    db_session.flush()
    _user(db_session, "hr", ["GROUP_HR"])
    headers = _token(client, "hr")

    generated = client.post(
        "/api/attendance-schedules/generate?period=2026-05",
        headers=headers,
    )

    assert generated.status_code == 200, generated.text
    record = db_session.scalars(
        select(AttendanceRecord).where(
            AttendanceRecord.employee_id == employee.id,
            AttendanceRecord.period == "2026-05",
        )
    ).one()
    assert record.expected_days == Decimal("22")
    assert record.generated_expected_days == Decimal("22")
    assert record.expected_days_rule_id == rule.id
    assert record.expected_days_adjust_reason is None


def test_schedule_generation_creates_first_attendance_record(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    employee = _emp(db_session, "E_NEW", gzs.id)
    employee.hire_date = date(2026, 5, 1)
    rule = _global_expected_days_rule(db_session, monthly_expected_days="21.5")
    _user(db_session, "hr-new-record", ["GROUP_HR"])
    headers = _token(client, "hr-new-record")

    generated = client.post(
        "/api/attendance-schedules/generate?period=2026-05",
        headers=headers,
    )

    assert generated.status_code == 200, generated.text
    assert generated.json() == {
        "period": "2026-05",
        "generated": 1,
        "adjusted_preserved": 0,
    }
    record = db_session.scalars(
        select(AttendanceRecord).where(
            AttendanceRecord.employee_id == employee.id,
            AttendanceRecord.period == "2026-05",
        )
    ).one()
    assert record.expected_days == Decimal("21.5")
    assert record.generated_expected_days == Decimal("21.5")
    assert record.expected_days_rule_id == rule.id


def test_schedule_generation_is_blocked_for_a_reopened_payroll_round(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    employee = _emp(db_session, "E_REOPENED", gzs.id)
    employee.hire_date = date(2026, 5, 1)
    rule = _global_expected_days_rule(db_session)
    record = AttendanceRecord(
        employee_id=employee.id,
        period="2026-05",
        generated_expected_days=Decimal("22"),
        expected_days_rule_id=rule.id,
        expected_days=Decimal("22"),
        actual_days=Decimal("20"),
    )
    db_session.add_all(
        [
            record,
            PayrollBatch(
                period="2026-05",
                attendance_start=date(2026, 5, 1),
                attendance_end=date(2026, 5, 31),
                status=BatchStatus.DRAFT,
                version=2,
            ),
        ]
    )
    rule.monthly_expected_days = Decimal("21")
    db_session.flush()
    _user(db_session, "hr-reopened-schedule", ["GROUP_HR"])

    generated = client.post(
        "/api/attendance-schedules/generate?period=2026-05",
        headers=_token(client, "hr-reopened-schedule"),
    )

    assert generated.status_code == 409
    db_session.refresh(record)
    assert record.generated_expected_days == Decimal("22")
    assert record.expected_days == Decimal("22")


def test_schedule_generation_only_targets_employees_active_in_the_period(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    active = _emp(db_session, "E_ACTIVE", gzs.id)
    active.hire_date = date(2026, 5, 1)
    future = _emp(db_session, "E_FUTURE", gzs.id)
    future.hire_date = date(2026, 6, 1)
    former = _emp(db_session, "E_FORMER", gzs.id)
    former.hire_date = date(2026, 1, 1)
    former.leave_date = date(2026, 4, 30)
    _global_expected_days_rule(db_session)
    _user(db_session, "hr-active-period", ["GROUP_HR"])

    generated = client.post(
        "/api/attendance-schedules/generate?period=2026-05",
        headers=_token(client, "hr-active-period"),
    )

    assert generated.status_code == 200, generated.text
    assert generated.json()["generated"] == 1
    employee_ids = set(
        db_session.scalars(
            select(AttendanceRecord.employee_id).where(AttendanceRecord.period == "2026-05")
        ).all()
    )
    assert employee_ids == {active.id}


def test_holiday_calendar_rejects_an_invalid_period_month(client, db_session):
    _user(db_session, "hr", ["GROUP_HR"])
    headers = _token(client, "hr")

    response = client.post(
        "/api/holiday-calendar/periods/2026-13/finalize",
        headers=headers,
    )

    assert response.status_code == 422


def test_store_manager_can_record_own_store(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    emp = _emp(db_session, "E1", gzs.id)
    _global_expected_days_rule(db_session)
    _user(db_session, "mgr", ["STORE_MANAGER"], scope_ids=[gzs.id])
    h = _token(client, "mgr")
    r = client.put(
        f"/api/employees/{emp.id}/attendance/2026-05",
        headers=h,
        json={"expected_days": "22", "actual_days": "22"},
    )
    assert r.status_code == 200


def test_store_manager_cannot_record_other_store(client, db_session):
    _gz, gzs, szs = _orgs(db_session)
    other = _emp(db_session, "SZ1", szs.id)
    _user(db_session, "mgr", ["STORE_MANAGER"], scope_ids=[gzs.id])
    h = _token(client, "mgr")
    r = client.put(
        f"/api/employees/{other.id}/attendance/2026-05",
        headers=h,
        json={"expected_days": "22", "actual_days": "22"},
    )
    assert r.status_code == 404  # 越权：他店员工不可见


def test_reopened_attendance_requires_the_payroll_correct_permission_scope(client, db_session):
    _gz, gzs, szs = _orgs(db_session)
    employee = _emp(db_session, "E_CORRECT_SCOPE", gzs.id)
    employee.department = Department.DINING
    record = AttendanceRecord(
        employee_id=employee.id,
        period="2026-05",
        expected_days=Decimal("22"),
        actual_days=Decimal("20"),
        worked_hours=Decimal("180"),
    )
    batch = PayrollBatch(
        period="2026-05",
        attendance_start=date(2026, 5, 1),
        attendance_end=date(2026, 5, 31),
        status=BatchStatus.DRAFT,
        version=2,
    )
    db_session.add_all([record, batch])
    db_session.flush()
    db_session.add(
        PayrollResult(
            batch_id=batch.id,
            employee_id=employee.id,
            batch_version=1,
            version=1,
            org_unit_id=gzs.id,
            department=Department.DINING,
            actual_attendance_days=Decimal("20"),
            gross=Decimal("5000"),
            deposit=Decimal("0"),
            net=Decimal("5000"),
            carry_forward=Decimal("0"),
            rule_version="v4",
            input_snapshot={},
            lines=[],
            exceptions=[],
            warnings=[],
            has_error=False,
        )
    )
    _mixed_attendance_correction_roles(db_session)
    _user(
        db_session,
        "mixed-attendance-corrector",
        ["GLOBAL_ATTENDANCE_ONLY", "SCOPED_PAYROLL_CORRECT_ONLY"],
        scope_ids=[szs.id],
    )
    db_session.flush()

    response = client.put(
        f"/api/employees/{employee.id}/attendance/2026-05",
        headers=_token(client, "mixed-attendance-corrector"),
        json={
            "expected_days": "22",
            "actual_days": "21",
            "worked_hours": "189",
            "correction_reason": "Approved correction",
            "attachment_url": "https://evidence.example/correction.pdf",
        },
    )

    assert response.status_code == 404
    db_session.refresh(record)
    assert record.actual_days == Decimal("20")
    assert record.worked_hours == Decimal("180")


def test_attendance_validation_rejects_out_of_range(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    emp = _emp(db_session, "E1", gzs.id)
    _user(db_session, "hr", ["GROUP_HR"])
    h = _token(client, "hr")
    r = client.put(
        f"/api/employees/{emp.id}/attendance/2026-05",
        headers=h,
        json={"expected_days": "22", "actual_days": "40"},  # >31
    )
    assert r.status_code == 422


def test_set_performance(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    emp = _emp(db_session, "E1", gzs.id)
    _user(db_session, "hr", ["GROUP_HR"])
    h = _token(client, "hr")
    r = client.put(
        f"/api/employees/{emp.id}/performance/2026-05",
        headers=h,
        json={"coefficient": "1.200", "score": "88"},
    )
    assert r.status_code == 200
    assert r.json()["coefficient"] == "1.200"
    audit_row = db_session.scalars(
        select(AuditLog).where(
            AuditLog.action == "performance.set",
            AuditLog.target_id == emp.id,
        )
    ).one()
    assert audit_row.detail["before"] == {"record_exists": False}
    assert audit_row.detail["after"] == {
        "record_exists": True,
        "coefficient": "1.20",
        "score": "88.00",
        "remark": None,
    }


def test_attendance_read_requires_permission(client, db_session):
    _orgs(db_session)
    _user(db_session, "emp", ["EMPLOYEE"])
    h = _token(client, "emp")
    assert client.get("/api/attendance?period=2026-05", headers=h).status_code == 403


def test_attendance_excel_import_scoped(client, db_session):
    _gz, gzs, szs = _orgs(db_session)
    e1 = _emp(db_session, "E1", gzs.id)  # noqa: F841 广州店，可见
    e2 = _emp(db_session, "E2", szs.id)  # noqa: F841 深圳店，越权
    _global_expected_days_rule(db_session)
    _user(db_session, "mgr", ["STORE_MANAGER"], scope_ids=[gzs.id])
    h = _token(client, "mgr")

    wb = Workbook()
    ws = wb.active
    ws.append(["工号", "姓名", "应出勤", "实出勤", "加班"])
    ws.append(["E1", "甲", 22, 21, 5])
    ws.append(["E2", "乙", 22, 22, 0])  # 越权工号
    ws.append(["E9", "丙", 22, 22, 0])  # 不存在
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    r = client.post(
        "/api/attendance/import?period=2026-05",
        headers=h,
        files={"file": ("att.xlsx", buf, "application/vnd.ms-excel")},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["matched"] == 1  # 仅广州店 E1
    assert set(body["skipped"]) == {"E2", "E9"}  # 越权 + 不存在都跳过


def test_attendance_import_rejects_duplicate_emp_no_across_sheets(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    _emp(db_session, "E1", gzs.id)
    _user(db_session, "hr", ["GROUP_HR"])
    headers = _token(client, "hr")

    workbook = Workbook()
    first = workbook.active
    first.title = "first"
    first.append(["工号", "应出勤", "实出勤"])
    first.append(["E1", 22, 22])
    second = workbook.create_sheet("second")
    second.append(["工号", "应出勤", "实出勤"])
    second.append(["E1", 21, 21])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = client.post(
        "/api/attendance/import?period=2026-05",
        headers=headers,
        files={"file": ("duplicate.xlsx", buffer, "application/vnd.ms-excel")},
    )

    assert response.status_code == 422
    assert "E1" in response.json()["detail"]
    assert client.get("/api/attendance?period=2026-05", headers=headers).json() == []


@pytest.mark.parametrize(
    ("header", "label"),
    [
        ("应出勤", "应出勤"),
        ("实出勤", "实出勤"),
        ("出勤工时", "出勤工时"),
        ("休息天数", "休息天数"),
        ("加班", "加班"),
        ("法定节假日出勤天数", "法定节假日出勤天数"),
        ("请假", "请假"),
    ],
)
def test_attendance_import_rejects_nonblank_malformed_numeric_cells(
    client, db_session, header, label
):
    _gz, gzs, _szs = _orgs(db_session)
    _emp(db_session, "E1", gzs.id)
    _global_expected_days_rule(db_session)
    _user(db_session, "hr-invalid-attendance", ["GROUP_HR"])
    headers = _token(client, "hr-invalid-attendance")

    values = {"工号": "E1", "应出勤": 22, "实出勤": 20, header: "not-a-number"}
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(list(values))
    sheet.append(list(values.values()))
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = client.post(
        "/api/attendance/import?period=2026-05",
        headers=headers,
        files={"file": ("invalid.xlsx", buffer, "application/vnd.ms-excel")},
    )

    assert response.status_code == 422
    assert label in response.json()["detail"]
    assert "格式无效" in response.json()["detail"]
    assert client.get("/api/attendance?period=2026-05", headers=headers).json() == []


def test_attendance_import_preserves_omitted_inputs_and_audits_expected_days(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    employee = _emp(db_session, "E1", gzs.id)
    _global_expected_days_rule(db_session)
    _user(db_session, "hr", ["GROUP_HR"])
    headers = _token(client, "hr")
    assert (
        client.put(
            f"/api/employees/{employee.id}/attendance/2026-05",
            headers=headers,
            json={
                "expected_days": "22",
                "actual_days": "20",
                "worked_hours": "160",
                "rest_days": "1",
                "holiday_worked_days": "2",
            },
        ).status_code
        == 200
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["工号", "应出勤", "实出勤", "应出勤调整原因"])
    sheet.append(["E1", 21, 21, "排班修正"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = client.post(
        "/api/attendance/import?period=2026-05",
        headers=headers,
        files={"file": ("att.xlsx", buffer, "application/vnd.ms-excel")},
    )
    assert response.status_code == 200, response.text
    record = client.get("/api/attendance?period=2026-05", headers=headers).json()[0]
    assert record["expected_days"] == "21.00"
    assert record["expected_days_adjust_reason"] == "排班修正"
    assert record["worked_hours"] == "160.00"
    assert record["rest_days"] == "1.00"
    assert record["holiday_worked_days"] == "2.00"
    audit_row = db_session.scalars(
        select(AuditLog)
        .where(
            AuditLog.action == "attendance.expected_days.adjust",
            AuditLog.target_id == employee.id,
        )
        .order_by(AuditLog.id.desc())
    ).first()
    assert audit_row is not None
    assert audit_row.detail["source"] == "attendance_import"
    assert audit_row.detail["before"]["expected_days"] == "22.00"
    assert audit_row.detail["after"]["expected_days"] == "21.00"
    import_row_audit = db_session.scalars(
        select(AuditLog).where(
            AuditLog.action == "attendance.import.row",
            AuditLog.target_id == employee.id,
        )
    ).one()
    assert import_row_audit.detail["before"]["actual_days"] == "20.00"
    assert import_row_audit.detail["after"]["actual_days"] == "21.00"
    assert import_row_audit.detail["before"]["worked_hours"] == "160.00"
    assert import_row_audit.detail["after"]["worked_hours"] == "160.00"


def test_attendance_import_requires_hours_for_new_hourly_employee(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    employee = _emp(db_session, "E1", gzs.id)
    employee.department = Department.DINING
    db_session.flush()
    _user(db_session, "hr", ["GROUP_HR"])
    headers = _token(client, "hr")

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["工号", "应出勤", "实出勤"])
    sheet.append(["E1", 22, 22])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = client.post(
        "/api/attendance/import?period=2026-05",
        headers=headers,
        files={"file": ("att.xlsx", buffer, "application/vnd.ms-excel")},
    )
    assert response.status_code == 422
    assert "工时" in response.json()["detail"]


def test_attendance_import_accepts_approved_days_for_a_named_special_position(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    employee = _emp(db_session, "E_DISH", gzs.id)
    employee.department = Department.KITCHEN
    employee.position_title = "洗碗岗位"
    employee.hire_date = date(2026, 5, 1)
    _global_expected_days_rule(db_session)
    db_session.flush()
    _user(db_session, "hr-special-import", ["GROUP_HR"])

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["工号", "应出勤", "实出勤"])
    sheet.append(["E_DISH", 22, 20])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = client.post(
        "/api/attendance/import?period=2026-05",
        headers=_token(client, "hr-special-import"),
        files={"file": ("att.xlsx", buffer, "application/vnd.ms-excel")},
    )

    assert response.status_code == 200, response.text
    record = db_session.scalars(
        select(AttendanceRecord).where(
            AttendanceRecord.employee_id == employee.id,
            AttendanceRecord.period == "2026-05",
        )
    ).one()
    assert record.actual_days == Decimal("20")
    assert record.worked_hours is None


def test_list_performance_filters_period_and_organization_scope(client, db_session):
    _gz, gzs, szs = _orgs(db_session)
    visible = _emp(db_session, "E1", gzs.id)
    hidden = _emp(db_session, "E2", szs.id)
    db_session.add_all(
        [
            PerformanceRecord(
                employee_id=visible.id,
                period="2026-05",
                coefficient=Decimal("1.100"),
                score=Decimal("91"),
                remark="visible",
            ),
            PerformanceRecord(
                employee_id=visible.id,
                period="2026-04",
                coefficient=Decimal("0.900"),
            ),
            PerformanceRecord(
                employee_id=hidden.id,
                period="2026-05",
                coefficient=Decimal("1.200"),
            ),
        ]
    )
    db_session.flush()
    _user(db_session, "mgr", ["STORE_MANAGER"], scope_ids=[gzs.id])

    response = client.get("/api/performance?period=2026-05", headers=_token(client, "mgr"))

    assert response.status_code == 200, response.text
    assert response.json() == [
        {
            "employee_id": visible.id,
            "period": "2026-05",
            "coefficient": "1.100",
            "score": "91.00",
            "remark": "visible",
        }
    ]


def test_list_performance_rejects_an_invalid_calendar_month(client, db_session):
    _user(db_session, "hr", ["GROUP_HR"])

    response = client.get("/api/performance?period=2026-13", headers=_token(client, "hr"))

    assert response.status_code == 422


def test_list_performance_requires_attendance_read_permission(client, db_session):
    _user(db_session, "employee", ["EMPLOYEE"])

    response = client.get("/api/performance?period=2026-05", headers=_token(client, "employee"))

    assert response.status_code == 403


def test_performance_excel_import_scopes_rows_defaults_new_records_and_preserves_omitted_fields(
    client, db_session
):
    _gz, gzs, szs = _orgs(db_session)
    existing_employee = _emp(db_session, "E1", gzs.id)
    new_employee = _emp(db_session, "E2", gzs.id)
    hidden_employee = _emp(db_session, "E3", szs.id)
    db_session.add(
        PerformanceRecord(
            employee_id=existing_employee.id,
            period="2026-05",
            coefficient=Decimal("0.900"),
            score=Decimal("88"),
            remark="keep existing details",
        )
    )
    db_session.flush()
    _user(db_session, "mgr", ["STORE_MANAGER"], scope_ids=[gzs.id])
    headers = _token(client, "mgr")

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["工号", "绩效系数", "绩效得分", "备注"])
    sheet.append(["E1", 1.2, None, None])
    sheet.append(["E2", None, 95, "new record"])
    sheet.append(["E3", 1.1, 90, "out of scope"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = client.post(
        "/api/performance/import?period=2026-05",
        headers=headers,
        files={"file": ("performance.xlsx", buffer, "application/vnd.ms-excel")},
    )

    assert response.status_code == 200, response.text
    assert response.json() == {"matched": 2, "skipped": ["E3"]}
    records = {
        record.employee_id: record
        for record in db_session.scalars(
            select(PerformanceRecord).where(PerformanceRecord.period == "2026-05")
        ).all()
    }
    assert records[existing_employee.id].coefficient == Decimal("1.200")
    assert records[existing_employee.id].score == Decimal("88")
    assert records[existing_employee.id].remark == "keep existing details"
    assert records[new_employee.id].coefficient == Decimal("1.000")
    assert records[new_employee.id].score == Decimal("95")
    assert records[new_employee.id].remark == "new record"
    assert hidden_employee.id not in records
    audit_row = db_session.scalars(
        select(AuditLog).where(AuditLog.action == "performance.import").order_by(AuditLog.id.desc())
    ).one()
    assert audit_row.detail == {"period": "2026-05", "matched": 2, "skipped": 1}
    row_audits = {
        row.target_id: row
        for row in db_session.scalars(
            select(AuditLog).where(AuditLog.action == "performance.import.row")
        ).all()
    }
    assert row_audits[existing_employee.id].detail["before"]["coefficient"] == "0.90"
    assert row_audits[existing_employee.id].detail["after"]["coefficient"] == "1.20"
    assert row_audits[new_employee.id].detail["before"] == {"record_exists": False}
    assert row_audits[new_employee.id].detail["after"]["score"] == "95.00"


def test_performance_excel_import_validates_every_row_before_writing(client, db_session):
    _gz, gzs, _szs = _orgs(db_session)
    existing_employee = _emp(db_session, "E1", gzs.id)
    invalid_employee = _emp(db_session, "E2", gzs.id)
    db_session.add(
        PerformanceRecord(
            employee_id=existing_employee.id,
            period="2026-05",
            coefficient=Decimal("0.900"),
            score=Decimal("80"),
            remark="original",
        )
    )
    db_session.flush()
    _user(db_session, "hr", ["GROUP_HR"])
    headers = _token(client, "hr")

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["工号", "绩效系数", "绩效得分", "备注"])
    sheet.append(["E1", 1.2, 90, "would be valid"])
    sheet.append(["E2", "not-a-number", 91, "invalid coefficient"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = client.post(
        "/api/performance/import?period=2026-05",
        headers=headers,
        files={"file": ("performance.xlsx", buffer, "application/vnd.ms-excel")},
    )

    assert response.status_code == 422
    assert "E2" in response.json()["detail"]
    existing = db_session.scalars(
        select(PerformanceRecord).where(
            PerformanceRecord.employee_id == existing_employee.id,
            PerformanceRecord.period == "2026-05",
        )
    ).one()
    assert existing.coefficient == Decimal("0.900")
    assert existing.score == Decimal("80")
    assert existing.remark == "original"
    assert (
        db_session.scalars(
            select(PerformanceRecord).where(
                PerformanceRecord.employee_id == invalid_employee.id,
                PerformanceRecord.period == "2026-05",
            )
        ).one_or_none()
        is None
    )


@pytest.mark.parametrize(
    ("period", "status", "version"),
    [
        ("2026-05", BatchStatus.LOCKED, 1),
        ("2026-06", BatchStatus.DRAFT, 2),
    ],
)
def test_performance_excel_import_is_blocked_for_locked_and_reopened_batches(
    client, db_session, period, status, version
):
    _gz, gzs, _szs = _orgs(db_session)
    employee = _emp(db_session, "E1", gzs.id)
    db_session.add(
        PayrollBatch(
            period=period,
            attendance_start=date.fromisoformat(f"{period}-01"),
            attendance_end=date.fromisoformat(f"{period}-28"),
            status=status,
            version=version,
        )
    )
    db_session.flush()
    _user(db_session, "hr", ["GROUP_HR"])
    headers = _token(client, "hr")

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["工号", "绩效系数"])
    sheet.append(["E1", 1.2])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = client.post(
        f"/api/performance/import?period={period}",
        headers=headers,
        files={"file": ("performance.xlsx", buffer, "application/vnd.ms-excel")},
    )

    assert response.status_code == 409
    assert (
        db_session.scalars(
            select(PerformanceRecord).where(
                PerformanceRecord.employee_id == employee.id,
                PerformanceRecord.period == period,
            )
        ).one_or_none()
        is None
    )


@pytest.mark.parametrize(
    ("period", "status", "version"),
    [
        ("2026-05", BatchStatus.LOCKED, 1),
        ("2026-06", BatchStatus.DRAFT, 2),
    ],
)
def test_set_performance_is_blocked_for_locked_and_reopened_batches(
    client, db_session, period, status, version
):
    _gz, gzs, _szs = _orgs(db_session)
    employee = _emp(db_session, "E1", gzs.id)
    db_session.add(
        PayrollBatch(
            period=period,
            attendance_start=date.fromisoformat(f"{period}-01"),
            attendance_end=date.fromisoformat(f"{period}-28"),
            status=status,
            version=version,
        )
    )
    db_session.flush()
    _user(db_session, "hr", ["GROUP_HR"])

    response = client.put(
        f"/api/employees/{employee.id}/performance/{period}",
        headers=_token(client, "hr"),
        json={"coefficient": "1.200", "score": "90"},
    )

    assert response.status_code == 409
    assert (
        db_session.scalars(
            select(PerformanceRecord).where(
                PerformanceRecord.employee_id == employee.id,
                PerformanceRecord.period == period,
            )
        ).one_or_none()
        is None
    )
