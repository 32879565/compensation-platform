import io
import re
import struct
from decimal import Decimal
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook

from app.importing import excel as excel_import
from app.importing.excel import read_salary_workbook
from app.importing.parser import (
    SalaryRow,
    auto_store_aliases,
    dedupe_rows,
    infer_month,
    is_shadow_row,
    normalize_header,
    normalize_store_name,
    parse_money,
    standard_store_name,
)
from app.routers.attendance import _parse_attendance_import_rows


def test_attendance_import_parser_rejects_duplicate_emp_no_across_sheets():
    workbook = Workbook()
    first = workbook.active
    first.append(["工号", "应出勤", "实出勤"])
    first.append(["E1", 22, 22])
    second = workbook.create_sheet("second")
    second.append(["工号", "应出勤", "实出勤"])
    second.append(["E1", 21, 21])

    with pytest.raises(ValueError, match="E1"):
        _parse_attendance_import_rows(workbook)


# ---------------- parse_money（不变量1/2：Decimal + 失败不归零）----------------
@pytest.mark.parametrize(
    "raw,expected",
    [
        (3500, Decimal("3500")),
        (3500.5, Decimal("3500.5")),
        (Decimal("3500.55"), Decimal("3500.55")),
        ("3500", Decimal("3500")),
        ("3,500.50", Decimal("3500.50")),
        ("3，500", Decimal("3500")),  # 全角逗号
        ("¥3500", Decimal("3500")),
        ("3500元", Decimal("3500")),
        ("(200)", Decimal("-200")),  # 会计括号负数
        ("-200", Decimal("-200")),
    ],
)
def test_parse_money_success(raw, expected):
    assert parse_money(raw) == expected


@pytest.mark.parametrize("raw", [None, "", "  ", "abc", "8%", "无", "-", "3500元整"])
def test_parse_money_failure_returns_none_not_zero(raw):
    # 关键修复：无法解析返回 None（上层报错），绝不返回 0
    assert parse_money(raw) is None


@pytest.mark.parametrize(
    "raw",
    [
        float("nan"),
        float("inf"),
        float("-inf"),
        Decimal("NaN"),
        Decimal("Infinity"),
        "NaN",
        "Infinity",
        "-Infinity",
    ],
)
def test_parse_money_rejects_non_finite_decimal_values(raw):
    assert parse_money(raw) is None


def test_parse_money_preserves_precision():
    assert parse_money("4500.55") == Decimal("4500.55")


# ---------------- normalize_header ----------------
def test_header_rename():
    assert normalize_header("职务") == "职位"
    assert normalize_header("底薪") == "综合薪资"
    assert normalize_header("矿工") == "旷工"  # 错别字修正
    assert normalize_header("通勤费") == "补贴"


def test_header_drop():
    assert normalize_header("序号") is None
    assert normalize_header("性别") is None
    assert normalize_header("2026-05-01") is None  # 日期形表头


def test_header_post_alias():
    assert normalize_header("传菜") == "传菜岗"
    assert normalize_header("经理") == "经理岗"


def test_header_month_salary_context():
    # N月综合薪资：仅当 N==文件月份才认作综合薪资
    assert normalize_header("5月综合薪资", month=5) == "综合薪资"
    assert normalize_header("4月综合薪资", month=5) is None


def test_header_legal_attendance_by_column():
    assert normalize_header("法定", column_index=45) == "法定出勤"
    assert normalize_header("法定", column_index=10) == "法定补贴"


def test_header_passthrough_unknown():
    assert normalize_header("综合薪资") == "综合薪资"
    assert normalize_header("某个未知字段") == "某个未知字段"


# ---------------- 门店归一化 ----------------
def test_store_alias_and_normalize():
    aliases = {"万科": "天河智慧城店"}
    assert standard_store_name("万科", aliases) == "天河智慧城店"
    assert normalize_store_name("海岸城店月工资表（C0.8）") == "海岸城店"


def test_auto_store_aliases():
    names = {"万科", "万科店", "广州店"}
    assert auto_store_aliases(names) == {"万科": "万科店"}


def test_workbook_applies_legacy_store_aliases_by_default():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "万科"
    sheet.append(["万科月工资表"])
    sheet.append(["工号", "姓名", "合计工资"])
    sheet.append(["E1", "张三", 5000])
    source = io.BytesIO()
    workbook.save(source)
    source.seek(0)

    result = read_salary_workbook(source, period="2026-05")

    assert result.rows[0].store_name == "天河智慧城店"


def test_workbook_skips_instruction_sheet_and_reads_template_data_sheet():
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "使用说明"
    instructions.append(["人事导入前请核对工号与门店"])
    sheet = workbook.create_sheet("海岸城店")
    sheet.append(["海岸城店"])
    sheet.append(["第 3 行为表头"])
    sheet.append(["工号", "姓名", "复核部门", "综合薪资", "实发工资"])
    sheet.append(["E1", "张三", "厅面", 5000, 4800])
    source = io.BytesIO()
    workbook.save(source)
    source.seek(0)

    result = read_salary_workbook(source, period="2026-07")

    assert result.warnings == []
    assert len(result.rows) == 1
    assert result.rows[0].emp_no == "E1"
    assert result.rows[0].store_name == "深圳海岸城店"
    assert result.rows[0].fields["复核部门"] == "厅面"
    assert result.rows[0].money["综合薪资"] == Decimal("5000")


def test_workbook_rejects_an_excessive_worksheet_column_dimension():
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=257, value="姓名")
    source = io.BytesIO()
    workbook.save(source)
    source.seek(0)

    with pytest.raises(excel_import.WorkbookLimitError, match="column"):
        read_salary_workbook(source, period="2026-07")


def test_workbook_does_not_silently_truncate_rows_when_dimension_is_underreported():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "广州店"
    sheet.append(["广州店"])
    sheet.append(["工号", "姓名", "复核部门", "实发工资"])
    sheet.append(["E1", "张三", "厅面", 4800])
    sheet.append(["E2", "李四", "厨房", 5200])
    original = io.BytesIO()
    workbook.save(original)
    original.seek(0)

    forged = io.BytesIO()
    with (
        ZipFile(original, mode="r") as source_archive,
        ZipFile(forged, mode="w", compression=ZIP_DEFLATED) as target_archive,
    ):
        for member in source_archive.infolist():
            content = source_archive.read(member.filename)
            if member.filename == "xl/worksheets/sheet1.xml":
                content = re.sub(
                    rb'<dimension ref="[^"]+"',
                    b'<dimension ref="A1:D3"',
                    content,
                    count=1,
                )
            target_archive.writestr(member, content)
    forged.seek(0)

    result = read_salary_workbook(forged, period="2026-07")

    assert [row.emp_no for row in result.rows] == ["E1", "E2"]


def test_workbook_rejects_an_excessive_archive_compression_ratio():
    source = io.BytesIO()
    with ZipFile(source, mode="w", compression=ZIP_DEFLATED) as archive:
        archive.writestr("xl/worksheets/sheet1.xml", b"0" * 1_000_000)
    source.seek(0)

    with pytest.raises(excel_import.WorkbookLimitError, match="compression ratio"):
        read_salary_workbook(source, period="2026-07")


def test_workbook_rejects_excessive_zip_entries_before_materializing_the_directory():
    # A valid-looking end-of-central-directory record is enough to prove the
    # bounded preflight happens before ZipFile creates one object per member.
    source = io.BytesIO(struct.pack("<4s4H2LH", b"PK\x05\x06", 0, 0, 1001, 1001, 0, 0, 0))

    with pytest.raises(excel_import.WorkbookLimitError, match="member count"):
        read_salary_workbook(source, period="2026-07")


def test_workbook_counts_central_directory_records_before_constructing_zipfile(monkeypatch):
    central_header = b"PK\x01\x02" + (b"\x00" * 42)
    central_directory = central_header * 1_001
    # The attacker lies in EOCD (declares one entry) while the directory has
    # 1,001 records. The bounded raw preflight must reject before ZipFile.
    eocd = struct.pack(
        "<4s4H2LH",
        b"PK\x05\x06",
        0,
        0,
        1,
        1,
        len(central_directory),
        0,
        0,
    )
    source = io.BytesIO(central_directory + eocd)
    monkeypatch.setattr(
        excel_import,
        "ZipFile",
        lambda *_args, **_kwargs: pytest.fail("ZipFile must not be constructed"),
    )

    with pytest.raises(excel_import.WorkbookLimitError, match="member count"):
        read_salary_workbook(source, period="2026-07")


def test_streamed_row_limits_do_not_trust_optional_worksheet_dimensions():
    with pytest.raises(excel_import.WorkbookLimitError, match="row dimension"):
        excel_import._validate_streamed_row(20_001, (), processed_cells=0)

    with pytest.raises(excel_import.WorkbookLimitError, match="column dimension"):
        excel_import._validate_streamed_row(1, (None,) * 257, processed_cells=0)

    with pytest.raises(excel_import.WorkbookLimitError, match="cell dimension"):
        excel_import._validate_streamed_row(
            1,
            (None,),
            processed_cells=excel_import.MAX_WORKSHEET_DIMENSION_CELLS,
        )


def test_workbook_rejects_an_excessive_number_of_worksheets():
    workbook = Workbook()
    for index in range(excel_import.MAX_WORKBOOK_SHEETS):
        workbook.create_sheet(f"S{index}")
    source = io.BytesIO()
    workbook.save(source)
    source.seek(0)

    with pytest.raises(excel_import.WorkbookLimitError, match="worksheet count"):
        read_salary_workbook(source, period="2026-07")


def test_workbook_global_cell_budget_is_cumulative_across_sheets():
    with pytest.raises(excel_import.WorkbookLimitError, match="workbook cell budget"):
        excel_import._accumulate_workbook_cells(
            excel_import.MAX_WORKBOOK_STREAMED_CELLS,
            additional_cells=1,
        )


def test_workbook_rejects_more_than_ten_thousand_employee_rows():
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("广州店")
    sheet.append(["工号", "姓名", "合计工资"])
    for row_number in range(10_001):
        sheet.append([f"E{row_number}", f"员工{row_number}", 5000])
    source = io.BytesIO()
    workbook.save(source)
    source.seek(0)

    with pytest.raises(ValueError, match="10000"):
        read_salary_workbook(source, period="2026-05")


@pytest.mark.parametrize(
    ("limit_name", "limit", "message"),
    [
        ("MAX_XLSX_ARCHIVE_MEMBERS", 1, "member"),
        ("MAX_XLSX_MEMBER_BYTES", 256, "member expanded"),
        ("MAX_XLSX_EXPANDED_BYTES", 1_024, "expanded"),
    ],
)
def test_workbook_rejects_xlsx_archives_over_safety_limits(monkeypatch, limit_name, limit, message):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "广州店"
    sheet.append(["工号", "姓名", "合计工资"])
    sheet.append(["E1", "张三", 5000])
    source = io.BytesIO()
    workbook.save(source)
    source.seek(0)
    monkeypatch.setattr(excel_import, limit_name, limit, raising=False)

    with pytest.raises(ValueError, match=message):
        read_salary_workbook(source, period="2026-05")


# ---------------- infer_month ----------------
def test_infer_month():
    assert infer_month("工资核对表/2026年/5月/深圳一区.xlsx") == "2026-05"
    assert infer_month("4月东莞.xlsx", default_month="2026-05") == "2026-04"


# ---------------- 去重（核心 bug 修复）----------------
def _row(name, store, period="2026-05", emp_no=None, **money):
    return SalaryRow(
        period=period,
        name=name,
        store_name=store,
        emp_no=emp_no,
        money={k: (Decimal(str(v)) if v is not None else None) for k, v in money.items()},
    )


def test_dedupe_same_name_different_store_both_kept():
    # 关键修复：两个不同门店的同名员工不再被误删
    rows = [
        _row("张伟", "广州店", 综合薪资=5000),
        _row("张伟", "深圳店", 综合薪资=6000),
    ]
    result = dedupe_rows(rows)
    assert len(result) == 2
    stores = {r.store_name for r in result}
    assert stores == {"广州店", "深圳店"}


def test_dedupe_single_zero_record_kept():
    # 关键修复：唯一一条记录即便看似影子行也保留（如全月请假实发0的员工）
    rows = [_row("李四", "广州店", 应发工资=0, 合计工资=0, 实发工资=0)]
    result = dedupe_rows(rows)
    assert len(result) == 1
    assert result[0].name == "李四"


def test_dedupe_multi_row_shadow_eliminated():
    # 同一人同店多条：影子行被淘汰，保留有金额的真实行
    rows = [
        _row("王五", "广州店", 应发工资=0),  # 影子占位
        _row("王五", "广州店", 综合薪资=5000, 合计工资=5000),
    ]
    result = dedupe_rows(rows)
    assert len(result) == 1
    assert result[0].money["综合薪资"] == Decimal("5000")


def test_dedupe_by_emp_no_when_present():
    # 有工号时按工号归并（同工号跨表重复取最优）
    rows = [
        _row("张三", "广州店", emp_no="E1", 合计工资=3000),
        _row("张三", "广州店", emp_no="E1", 合计工资=5000),
    ]
    result = dedupe_rows(rows)
    assert len(result) == 1
    assert result[0].money["合计工资"] == Decimal("5000")


def test_is_shadow_row():
    assert is_shadow_row(_row("A", "店", 应发工资=0, 合计工资=0)) is True
    assert is_shadow_row(_row("A", "店", 综合薪资=5000)) is False


# ---------------- 复核修复：占位工号不作身份键 ----------------
def test_placeholder_emp_no_normalized_to_none():
    from app.importing.parser import normalize_emp_no

    for token in ("无", "-", "／", "N/A", "0", "", "  "):
        assert normalize_emp_no(token) is None
    assert normalize_emp_no("E1001") == "E1001"


def test_placeholder_emp_no_does_not_collapse_distinct_people():
    # 三个不同人工号都填占位符"无"→ 归一为 None → 按 (姓名,门店) 区分，不被合并
    rows = [
        _row("张三", "广州店", emp_no=None, 合计工资=5000),
        _row("李四", "广州店", emp_no=None, 合计工资=6000),
        _row("王五", "广州店", emp_no=None, 合计工资=7000),
    ]
    assert len(dedupe_rows(rows)) == 3


def test_same_emp_no_cross_store_kept_separate():
    # 员工工号全局唯一；当前导入的身份键必须始终是 (周期, 工号)，不能把门店
    # 混入键中而让同一员工在同一周期写入两份工资记录。
    rows = [
        _row("张三", "广州店", emp_no="E1", 合计工资=3000),
        _row("张三", "深圳店", emp_no="E1", 合计工资=2500),
    ]
    result = dedupe_rows(rows)
    assert len(result) == 1
    assert result[0].identity_key() == ("2026-05", "E1")


def test_dedupe_does_not_use_name_as_an_identity_when_emp_no_is_missing():
    # 无工号的当前行会在暂存校验阶段变成可操作的错误；这里不能按姓名静默合并，
    # 否则人工认领时会丢失原始行。
    rows = [
        _row("张三", "广州店", 合计工资=3000),
        _row("张三", "广州店", 合计工资=2500),
    ]

    assert len(dedupe_rows(rows)) == 2


def test_full_width_minus_parsed():
    assert parse_money("－150") == Decimal("-150")


# ---------------- 复核修复：has_month_salary 丢弃普通综合薪资 ----------------
def test_has_month_salary_drops_plain_comprehensive():
    # 存在"5月综合薪资"当月列时，普通"综合薪资"/"底薪"列应丢弃避免碰撞
    assert normalize_header("综合薪资", month=5, has_month_salary=True) is None
    assert normalize_header("底薪", month=5, has_month_salary=True) is None
    assert normalize_header("5月综合薪资", month=5, has_month_salary=True) == "综合薪资"
    # 无当月列时普通综合薪资照常保留
    assert normalize_header("综合薪资", month=5, has_month_salary=False) == "综合薪资"


def test_numeric_header_dropped():
    assert normalize_header("3000") is None
    assert normalize_header("-150.5") is None
