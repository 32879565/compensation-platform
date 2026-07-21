import io
from decimal import Decimal

import pytest
from openpyxl import Workbook

from app.importing.excel import read_salary_workbook
from app.importing.parser import SalaryRow
from app.importing.service import ImportError_, confirm_import, stage_import
from app.models.employee import Employee
from app.models.org import OrgType, OrgUnit
from app.models.salary import ImportStagingRow, ImportStatus, RowStatus, SalaryRecord, SalarySource

pytestmark = pytest.mark.usefixtures("pg_engine")


def _row(name, store, period="2026-05", emp_no=None, fields=None, money=None):
    return SalaryRow(
        period=period,
        name=name,
        store_name=store,
        emp_no=emp_no,
        fields=fields or {},
        money=money or {},
    )


def _employee(session, store, emp_no, name):
    employee = Employee(emp_no=emp_no, name=name, org_unit_id=store.id)
    session.add(employee)
    session.flush()
    return employee


def test_stage_flags_unparseable_money_as_error(db_session):
    store = OrgUnit(code="S1", name="广州店", type=OrgType.STORE, city="广州")
    db_session.add(store)
    db_session.flush()
    _employee(db_session, store, "E1", "张三")
    _employee(db_session, store, "E2", "李四")
    rows = [
        _row(
            "张三",
            "广州店",
            emp_no="E1",
            fields={"合计工资": "5000"},
            money={"合计工资": Decimal("5000")},
        ),
        # 有原值文本但无法解析 → 应标记为 ERROR（不静默归零）
        _row(
            "李四",
            "广州店",
            emp_no="E2",
            fields={"合计工资": "五千元"},
            money={"合计工资": None},
        ),
    ]
    batch = stage_import(db_session, filename="t.xlsx", period="2026-05", rows=rows)
    assert batch.total_rows == 2
    assert batch.error_rows == 1


def test_confirm_blocked_when_errors_present(db_session):
    store = OrgUnit(code="S1", name="广州店", type=OrgType.STORE, city="广州")
    db_session.add(store)
    db_session.flush()
    _employee(db_session, store, "E2", "李四")
    rows = [
        _row(
            "李四",
            "广州店",
            emp_no="E2",
            fields={"合计工资": "五千"},
            money={"合计工资": None},
        )
    ]
    batch = stage_import(db_session, filename="t.xlsx", period="2026-05", rows=rows)
    with pytest.raises(ImportError_):
        confirm_import(db_session, batch)


def test_confirm_writes_records_and_resolves_org(db_session):
    store = OrgUnit(code="S1", name="广州店", type=OrgType.STORE, city="广州")
    db_session.add(store)
    db_session.flush()
    zhang_employee = _employee(db_session, store, "E1", "张三")
    _employee(db_session, store, "E2", "王五")
    rows = [
        _row(
            "张三",
            "广州店",
            emp_no="E1",
            fields={"合计工资": "5000"},
            money={"合计工资": Decimal("5000")},
        ),
        _row(
            "王五",
            "广州店",
            emp_no="E2",
            fields={"合计工资": "6000"},
            money={"合计工资": Decimal("6000")},
        ),
    ]
    batch = stage_import(db_session, filename="t.xlsx", period="2026-05", rows=rows)
    written = confirm_import(db_session, batch)
    assert written == 2
    assert batch.status == ImportStatus.CONFIRMED

    recs = db_session.query(SalaryRecord).all()
    assert len(recs) == 2
    zhang = next(r for r in recs if r.name == "张三")
    assert zhang.org_unit_id == store.id  # 门店名解析到组织
    assert zhang.fields["合计工资"] == "5000"  # 金额以字符串保精度
    assert zhang.source == SalarySource.IMPORT
    assert zhang.employee_id == zhang_employee.id


def test_current_import_rejects_employee_master_store_mismatch(db_session):
    source_store = OrgUnit(code="S1", name="Source Store", type=OrgType.STORE, city="Guangzhou")
    workbook_store = OrgUnit(code="S2", name="Workbook Store", type=OrgType.STORE, city="Shenzhen")
    db_session.add_all([source_store, workbook_store])
    db_session.flush()
    _employee(db_session, source_store, "E1", "Employee One")
    batch = stage_import(
        db_session,
        filename="wrong-store.xlsx",
        period="2026-05",
        rows=[
            _row(
                "Employee One",
                "Workbook Store",
                emp_no="E1",
                fields={"合计工资": "5000"},
                money={"合计工资": Decimal("5000")},
            )
        ],
    )

    staged = db_session.query(ImportStagingRow).filter_by(batch_id=batch.id).one()
    assert staged.status == RowStatus.ERROR
    assert any("master store" in error.lower() for error in staged.errors)
    with pytest.raises(ImportError_):
        confirm_import(db_session, batch)
    assert db_session.query(SalaryRecord).count() == 0


def test_confirm_rechecks_employee_master_store_after_staging(db_session):
    staged_store = OrgUnit(code="S1", name="Staged Store", type=OrgType.STORE, city="Guangzhou")
    transferred_store = OrgUnit(
        code="S2", name="Transferred Store", type=OrgType.STORE, city="Shenzhen"
    )
    db_session.add_all([staged_store, transferred_store])
    db_session.flush()
    employee = _employee(db_session, staged_store, "E1", "Employee One")
    batch = stage_import(
        db_session,
        filename="staged-before-transfer.xlsx",
        period="2026-05",
        rows=[
            _row(
                "Employee One",
                "Staged Store",
                emp_no="E1",
                fields={"合计工资": "5000"},
                money={"合计工资": Decimal("5000")},
            )
        ],
    )
    assert batch.error_rows == 0
    employee.org_unit_id = transferred_store.id
    db_session.flush()

    with pytest.raises(ImportError_, match="does not match the employee master store"):
        confirm_import(db_session, batch)

    assert db_session.query(SalaryRecord).count() == 0


def test_confirm_twice_rejected(db_session):
    store = OrgUnit(code="S1", name="广州店", type=OrgType.STORE, city="广州")
    db_session.add(store)
    db_session.flush()
    _employee(db_session, store, "E1", "张三")
    rows = [
        _row(
            "张三",
            "广州店",
            emp_no="E1",
            fields={"合计工资": "5000"},
            money={"合计工资": Decimal("5000")},
        )
    ]
    batch = stage_import(db_session, filename="t.xlsx", period="2026-05", rows=rows)
    confirm_import(db_session, batch)
    with pytest.raises(ImportError_):
        confirm_import(db_session, batch)


def test_staging_row_status(db_session):
    store = OrgUnit(code="S1", name="广州店", type=OrgType.STORE, city="广州")
    db_session.add(store)
    db_session.flush()
    _employee(db_session, store, "E1", "张三")
    rows = [
        _row(
            "张三",
            "广州店",
            emp_no="E1",
            fields={"合计工资": "5000"},
            money={"合计工资": Decimal("5000")},
        )
    ]
    batch = stage_import(db_session, filename="t.xlsx", period="2026-05", rows=rows)
    from app.models.salary import ImportStagingRow

    srows = db_session.query(ImportStagingRow).filter_by(batch_id=batch.id).all()
    assert len(srows) == 1
    assert srows[0].status == RowStatus.OK


# ---------------- Excel 端到端读取 ----------------
def _make_workbook() -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "广州店"
    ws.append(["广州店月工资表"])  # 首行门店标题
    ws.append(["工号", "姓名", "综合薪资", "合计工资", "实发工资"])  # 表头
    ws.append(["E1", "张三", 5000, 5200, 4800])
    ws.append(["E2", "李四", 6000, 6200, 5800])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_read_salary_workbook(db_session):
    result = read_salary_workbook(_make_workbook(), period="2026-05")
    assert len(result.rows) == 2
    zhang = next(r for r in result.rows if r.name == "张三")
    assert zhang.store_name == "广州店"
    assert zhang.money["综合薪资"] == Decimal("5000")
    assert zhang.money["合计工资"] == Decimal("5200")


def _make_collision_workbook() -> io.BytesIO:
    # 同名金额列碰撞 + SUM 字段跨列
    wb = Workbook()
    ws = wb.active
    ws.title = "广州店"
    ws.append(["广州店"])
    # 底薪(→综合薪资) 与 现综合薪资(→综合薪资，空) 碰撞；元旦法定+清明法定→法定出勤求和
    ws.append(["姓名", "底薪", "现综合薪资", "元旦法定", "清明法定"])
    ws.append(["张三", 5000, None, 1, 1])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_empty_later_column_does_not_wipe_money(db_session):
    # 复核修复：靠右的空"现综合薪资"列不得把"底薪"解析出的 5000 覆盖成 None
    result = read_salary_workbook(_make_collision_workbook(), period="2026-05")
    assert len(result.rows) == 1
    row = result.rows[0]
    assert row.money["综合薪资"] == Decimal("5000")


def test_sum_merged_field_accumulates(db_session):
    # 复核修复：元旦法定+清明法定 跨列累加为法定出勤=2，而非覆盖成 1
    result = read_salary_workbook(_make_collision_workbook(), period="2026-05")
    assert result.rows[0].money["法定出勤"] == Decimal("2")


def test_summary_row_and_bad_sheet_skipped(db_session):
    wb = Workbook()
    ws = wb.active
    ws.title = "广州店"
    ws.append(["广州店"])
    ws.append(["姓名", "合计工资"])
    ws.append(["张三", 5000])
    ws.append(["合计", 5000])  # 表尾汇总行，不应成为员工
    extra = wb.create_sheet("调店")  # 非员工工作表，应跳过
    extra.append(["姓名", "合计工资"])
    extra.append(["李四", 9999])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    result = read_salary_workbook(buf, period="2026-05")
    names = {r.name for r in result.rows}
    assert names == {"张三"}  # 合计行与调店表都被排除


def test_read_then_stage_then_confirm_e2e(db_session):
    store = OrgUnit(code="S1", name="广州店", type=OrgType.STORE, city="广州")
    db_session.add(store)
    db_session.flush()
    _employee(db_session, store, "E1", "张三")
    _employee(db_session, store, "E2", "李四")
    result = read_salary_workbook(_make_workbook(), period="2026-05")
    batch = stage_import(db_session, filename="广州店.xlsx", period="2026-05", rows=result.rows)
    assert batch.error_rows == 0
    written = confirm_import(db_session, batch)
    assert written == 2
    assert db_session.query(SalaryRecord).count() == 2


def test_current_import_stages_missing_emp_no_for_manual_claim(db_session):
    batch = stage_import(
        db_session,
        filename="missing-emp-no.xlsx",
        period="2026-05",
        rows=[
            _row("张三", "广州店", fields={"合计工资": "5000"}, money={"合计工资": Decimal("5000")})
        ],
    )

    staged = db_session.query(ImportStagingRow).filter_by(batch_id=batch.id).one()
    assert batch.error_rows == 1
    assert staged.status == RowStatus.ERROR
    assert any("缺少工号" in error and "人工认领" in error for error in staged.errors)


def test_current_import_stages_unknown_emp_no_for_manual_claim(db_session):
    batch = stage_import(
        db_session,
        filename="unknown-emp-no.xlsx",
        period="2026-05",
        rows=[
            _row(
                "张三",
                "广州店",
                emp_no="UNKNOWN",
                fields={"合计工资": "5000"},
                money={"合计工资": Decimal("5000")},
            )
        ],
    )

    staged = db_session.query(ImportStagingRow).filter_by(batch_id=batch.id).one()
    assert batch.error_rows == 1
    assert staged.status == RowStatus.ERROR
    assert any("UNKNOWN" in error and "人工认领" in error for error in staged.errors)


def test_historical_staging_preserves_rows_without_emp_no(db_session):
    batch = stage_import(
        db_session,
        filename="historical.xlsx",
        period="2026-05",
        source=SalarySource.HISTORICAL,
        rows=[
            _row("张三", "广州店", fields={"合计工资": "5000"}, money={"合计工资": Decimal("5000")})
        ],
    )

    assert batch.error_rows == 0


def test_current_import_stages_conflicting_store_for_same_period_emp_no(db_session):
    guangzhou = OrgUnit(code="S1", name="广州店", type=OrgType.STORE, city="广州")
    shenzhen = OrgUnit(code="S2", name="深圳店", type=OrgType.STORE, city="深圳")
    db_session.add_all([guangzhou, shenzhen])
    db_session.flush()
    _employee(db_session, guangzhou, "E1", "张三")
    batch = stage_import(
        db_session,
        filename="conflicting-identity.xlsx",
        period="2026-05",
        rows=[
            _row(
                "张三",
                "广州店",
                emp_no="E1",
                fields={"合计工资": "5000"},
                money={"合计工资": Decimal("5000")},
            ),
            _row(
                "张三",
                "深圳店",
                emp_no="E1",
                fields={"合计工资": "5000"},
                money={"合计工资": Decimal("5000")},
            ),
        ],
    )

    staged = db_session.query(ImportStagingRow).filter_by(batch_id=batch.id).one()
    assert batch.total_rows == 1
    assert staged.status == RowStatus.ERROR
    assert any("多个姓名或门店" in error for error in staged.errors)
