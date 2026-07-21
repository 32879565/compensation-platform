from __future__ import annotations

import io
from dataclasses import dataclass
from decimal import Decimal
from typing import Literal, overload

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel, Field, ValidationError, field_validator
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.attendance.schedule import ExpectedDaysError, resolve_expected_days
from app.audit import service as audit
from app.auth.deps import require_permission
from app.auth.permissions import Perm
from app.auth.service import (
    Principal,
    permission_org_scope_allows,
    resolve_permission_org_scope,
)
from app.core.decimal import decimal_text
from app.core.urls import optional_http_url
from app.db.session import get_session
from app.importing.parser import clean_text, parse_money
from app.models.attendance import AttendanceRecord, PerformanceRecord
from app.models.employee import Employee, requires_approved_attendance_days
from app.models.payroll_batch import PayrollBatch
from app.models.payroll_result import AdjustmentRecord, PayrollResult
from app.payroll.batch_service import allowed_attendance_fields
from app.payroll.guards import PayrollSourceLockedError, assert_period_mutable
from app.repositories.employee import EmployeeRepository

router = APIRouter(prefix="/api", tags=["attendance"])

_PERIOD = r"^\d{4}-\d{2}$"


def _visible_employee(
    session: Session, org_scope: frozenset[int] | None, employee_id: int
) -> Employee:
    emp = EmployeeRepository(session, org_scope=org_scope).get(employee_id)
    if emp is None:
        raise HTTPException(status_code=404, detail="员工不存在或不可见")
    return emp


def _ensure_period_mutable(session: Session, period: str) -> bool:
    try:
        correction_round = assert_period_mutable(session, period)
    except PayrollSourceLockedError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from None
    return correction_round


# ------------------- 考勤 -------------------
class AttendanceBody(BaseModel):
    expected_days: Decimal = Field(ge=0, le=31, max_digits=6, decimal_places=2)
    expected_days_adjust_reason: str | None = Field(default=None, max_length=255)
    actual_days: Decimal = Field(ge=0, le=31, max_digits=6, decimal_places=2)
    worked_hours: Decimal | None = Field(default=None, ge=0, le=744, max_digits=6, decimal_places=2)
    rest_days: Decimal = Field(default=Decimal(0), ge=0, le=31, max_digits=6, decimal_places=2)
    overtime_hours: Decimal = Field(
        default=Decimal(0), ge=0, le=744, max_digits=6, decimal_places=2
    )
    holiday_worked_days: Decimal = Field(
        default=Decimal(0), ge=0, le=31, max_digits=6, decimal_places=2
    )
    leave_days: Decimal = Field(default=Decimal(0), ge=0, le=31, max_digits=6, decimal_places=2)
    late_count: int = Field(default=0, ge=0)
    early_leave_count: int = Field(default=0, ge=0)
    correction_reason: str | None = Field(default=None, max_length=1000)
    attachment_url: str | None = Field(default=None, max_length=512)

    @field_validator("expected_days_adjust_reason", "correction_reason", mode="before")
    @classmethod
    def strip_optional_text(cls, value: object) -> object:
        """Normalize blank audit fields to absent values before mutable-state checks."""
        if isinstance(value, str):
            return value.strip() or None
        return value

    @field_validator("attachment_url", mode="before")
    @classmethod
    def validate_attachment_url(cls, value: object) -> object:
        return optional_http_url(value)


class AttendanceOut(BaseModel):
    employee_id: int
    period: str
    generated_expected_days: Decimal | None
    expected_days_rule_id: int | None
    expected_days: Decimal
    expected_days_adjust_reason: str | None
    actual_days: Decimal
    worked_hours: Decimal | None
    rest_days: Decimal
    overtime_hours: Decimal
    holiday_worked_days: Decimal
    leave_days: Decimal
    late_count: int
    early_leave_count: int

    model_config = {"from_attributes": True}


@dataclass(frozen=True)
class _ExpectedDaysResolution:
    """The generated baseline plus any explicitly authorized HR exception."""

    body: AttendanceBody
    generated_expected_days: Decimal | None
    expected_days_rule_id: int | None
    audit_reason: str | None


def _require_expected_days_adjust_permission(
    session: Session, principal: Principal, employee: Employee
) -> None:
    """Enforce the separate HR authorization for a payroll-basis exception."""
    if not principal.has_permission(Perm.ATTENDANCE_EXPECTED_DAYS_ADJUST):
        raise HTTPException(
            status_code=403,
            detail="Only Group HR may adjust generated expected days.",
        )
    scope = resolve_permission_org_scope(session, principal, Perm.ATTENDANCE_EXPECTED_DAYS_ADJUST)
    if scope is not None and employee.org_unit_id not in scope:
        raise HTTPException(
            status_code=403,
            detail="The expected-days adjustment is outside your scope.",
        )


def _resolve_expected_days_for_write(
    session: Session,
    *,
    employee: Employee,
    period: str,
    body: AttendanceBody,
    existing: AttendanceRecord | None,
    principal: Principal,
) -> _ExpectedDaysResolution:
    """Keep a persisted payroll basis stable until an explicit HR action changes it.

    Normal attendance writes update actual attendance only.  They must not
    resolve a current schedule rule again: doing so would make a later rule
    edit silently rewrite an already-entered payroll basis.  The dedicated
    schedule-generation endpoint is the only rule-rebase operation.
    """
    requested_reason = body.expected_days_adjust_reason
    if existing is not None:
        generated_days = existing.generated_expected_days
        rule_id = existing.expected_days_rule_id

        if requested_reason is None:
            return _ExpectedDaysResolution(
                body=body.model_copy(
                    update={
                        "expected_days": existing.expected_days,
                        "expected_days_adjust_reason": existing.expected_days_adjust_reason,
                    }
                ),
                generated_expected_days=generated_days,
                expected_days_rule_id=rule_id,
                audit_reason=None,
            )

        if (
            body.expected_days == existing.expected_days
            and requested_reason == existing.expected_days_adjust_reason
        ):
            return _ExpectedDaysResolution(
                body=body.model_copy(
                    update={"expected_days_adjust_reason": existing.expected_days_adjust_reason}
                ),
                generated_expected_days=generated_days,
                expected_days_rule_id=rule_id,
                audit_reason=None,
            )

        # A legacy reason predates the dedicated HR permission and is not a
        # trustworthy authorization record. Generation establishes an
        # auditable baseline first; HR can then add a fresh exception.
        if generated_days is None or rule_id is None:
            _require_expected_days_adjust_permission(session, principal, employee)
            raise HTTPException(
                status_code=422,
                detail="Generate expected days before adjusting a legacy attendance record.",
            )

        _require_expected_days_adjust_permission(session, principal, employee)
        if body.expected_days == existing.expected_days:
            raise HTTPException(
                status_code=422,
                detail="Expected days did not change; use schedule generation to rebase the rule.",
            )
        if requested_reason == existing.expected_days_adjust_reason:
            raise HTTPException(
                status_code=422,
                detail="Changing expected days requires a new adjustment reason.",
            )

        # Returning to the stored generated baseline clears a prior exception.
        # A new rule is applied only through the audited generation endpoint.
        persisted_reason = requested_reason if body.expected_days != generated_days else None
        return _ExpectedDaysResolution(
            body=body.model_copy(update={"expected_days_adjust_reason": persisted_reason}),
            generated_expected_days=generated_days,
            expected_days_rule_id=rule_id,
            audit_reason=requested_reason,
        )

    try:
        generated = resolve_expected_days(session, employee, period)
    except ExpectedDaysError:
        if requested_reason:
            _require_expected_days_adjust_permission(session, principal, employee)
        raise HTTPException(
            status_code=422,
            detail="Create an expected-days schedule before recording attendance.",
        ) from None

    if requested_reason is None:
        return _ExpectedDaysResolution(
            body=body.model_copy(
                update={
                    "expected_days": generated.days,
                    "expected_days_adjust_reason": None,
                }
            ),
            generated_expected_days=generated.days,
            expected_days_rule_id=generated.rule_id,
            audit_reason=None,
        )

    _require_expected_days_adjust_permission(session, principal, employee)
    persisted_reason = requested_reason if body.expected_days != generated.days else None
    return _ExpectedDaysResolution(
        body=body.model_copy(update={"expected_days_adjust_reason": persisted_reason}),
        generated_expected_days=generated.days,
        expected_days_rule_id=generated.rule_id,
        audit_reason=requested_reason,
    )


def _upsert_attendance(
    session: Session,
    employee_id: int,
    period: str,
    body: AttendanceBody,
    *,
    generated_expected_days: Decimal | None = None,
    expected_days_rule_id: int | None = None,
    existing: AttendanceRecord | None = None,
    already_loaded: bool = False,
    flush: bool = True,
) -> AttendanceRecord:
    values = body.model_dump(
        exclude={"correction_reason", "attachment_url", "expected_days_adjust_reason"}
    )
    rec = existing
    if not already_loaded:
        rec = session.scalars(
            select(AttendanceRecord).where(
                AttendanceRecord.employee_id == employee_id, AttendanceRecord.period == period
            )
        ).first()
    if rec is None:
        rec = AttendanceRecord(
            employee_id=employee_id,
            period=period,
            generated_expected_days=generated_expected_days,
            expected_days_rule_id=expected_days_rule_id,
            expected_days_adjust_reason=body.expected_days_adjust_reason,
            **values,
        )
        session.add(rec)
    else:
        expected_days_changed = values["expected_days"] != rec.expected_days
        for field, value in values.items():
            setattr(rec, field, value)
        rec.generated_expected_days = generated_expected_days
        rec.expected_days_rule_id = expected_days_rule_id
        if expected_days_changed:
            rec.expected_days_adjust_reason = body.expected_days_adjust_reason
    if flush:
        session.flush()
    return rec


def _attendance_snapshot(rec: AttendanceRecord | None) -> dict:
    if rec is None:
        return {"record_exists": False}
    return {
        "record_exists": True,
        "generated_expected_days": decimal_text(rec.generated_expected_days),
        "expected_days_rule_id": rec.expected_days_rule_id,
        "expected_days": decimal_text(rec.expected_days),
        "expected_days_adjust_reason": rec.expected_days_adjust_reason,
        "actual_days": decimal_text(rec.actual_days),
        "worked_hours": decimal_text(rec.worked_hours),
        "rest_days": decimal_text(rec.rest_days),
        "overtime_hours": decimal_text(rec.overtime_hours),
        "holiday_worked_days": decimal_text(rec.holiday_worked_days),
        "leave_days": decimal_text(rec.leave_days),
        "late_count": rec.late_count,
        "early_leave_count": rec.early_leave_count,
    }


def _attendance_calculation_values(snapshot: dict) -> tuple[object, ...]:
    if not snapshot.get("record_exists"):
        return (False,)
    return tuple(
        snapshot[field]
        for field in (
            "expected_days",
            "actual_days",
            "worked_hours",
            "rest_days",
            "overtime_hours",
            "holiday_worked_days",
        )
    )


def _attendance_snapshot_from_body(body: AttendanceBody) -> dict:
    return {
        "record_exists": True,
        "expected_days": decimal_text(body.expected_days),
        "actual_days": decimal_text(body.actual_days),
        "worked_hours": decimal_text(body.worked_hours),
        "rest_days": decimal_text(body.rest_days),
        "overtime_hours": decimal_text(body.overtime_hours),
        "holiday_worked_days": decimal_text(body.holiday_worked_days),
    }


def _record_reopened_source_correction(
    session: Session,
    *,
    batch: PayrollBatch,
    employee_id: int,
    item: str,
    before: dict,
    after: dict,
    reason: str,
    attachment_url: str | None,
    principal: Principal,
) -> None:
    session.add(
        AdjustmentRecord(
            batch_id=batch.id,
            batch_version=batch.version,
            employee_id=employee_id,
            dispute_id=None,
            item=item,
            before_value=before,
            after_value=after,
            reason=reason,
            applicant_id=principal.user_id,
            approver_id=principal.user_id,
            attachment_url=attachment_url,
            recompute_result={
                "status": "PENDING_RERUN",
                "batch_version": batch.version,
            },
        )
    )


def _reopened_batch_employee_or_error(
    session: Session, period: str, employee_id: int
) -> tuple[PayrollBatch, PayrollResult]:
    """Ensure a direct correction can produce a replacement result.

    A reopened round has no active result yet, so the employee must have a
    result from a prior round of this batch.  Without this check a future hire
    could create a pending correction that every rerun fails to reconcile.
    """
    batch = session.scalars(select(PayrollBatch).where(PayrollBatch.period == period)).one()
    prior_result = session.scalars(
        select(PayrollResult)
        .where(
            PayrollResult.batch_id == batch.id,
            PayrollResult.employee_id == employee_id,
        )
        .order_by(PayrollResult.batch_version.desc(), PayrollResult.version.desc())
        .limit(1)
    ).first()
    if prior_result is None:
        raise HTTPException(
            status_code=422,
            detail="该员工不在已解锁薪资批次的原始核算范围内，不能发起更正",
        )
    return batch, prior_result


@router.put("/employees/{employee_id}/attendance/{period}", response_model=AttendanceOut)
def set_attendance(
    employee_id: int,
    period: str,
    body: AttendanceBody,
    principal: Principal = Depends(require_permission(Perm.ATTENDANCE_WRITE)),
    session: Session = Depends(get_session),
) -> AttendanceRecord:
    if not _period_ok(period):
        raise HTTPException(status_code=422, detail="周期格式应为 YYYY-MM")
    correction_round = _ensure_period_mutable(session, period)
    employee = _visible_employee(
        session,
        resolve_permission_org_scope(session, principal, Perm.ATTENDANCE_WRITE),
        employee_id,
    )
    correction_batch: PayrollBatch | None = None
    correction_prior_result: PayrollResult | None = None
    if correction_round:
        correction_batch, correction_prior_result = _reopened_batch_employee_or_error(
            session, period, employee_id
        )
        if not permission_org_scope_allows(
            session,
            principal,
            Perm.PAYROLL_CORRECT,
            correction_prior_result.org_unit_id,
        ):
            raise HTTPException(status_code=404, detail="员工不存在或不可见")
    existing = session.scalars(
        select(AttendanceRecord).where(
            AttendanceRecord.employee_id == employee_id,
            AttendanceRecord.period == period,
        )
    ).first()
    expected_resolution = _resolve_expected_days_for_write(
        session,
        employee=employee,
        period=period,
        body=body,
        existing=existing,
        principal=principal,
    )
    body = expected_resolution.body
    expected_days_adjusted = expected_resolution.audit_reason is not None
    if correction_round and not body.correction_reason:
        raise HTTPException(status_code=422, detail="更正已解锁批次的源数据必须填写更正原因")
    if correction_round and not body.attachment_url:
        raise HTTPException(status_code=422, detail="更正已解锁批次的源数据必须上传证明附件")
    before = _attendance_snapshot(existing)
    if correction_round:
        if correction_prior_result is None:
            raise HTTPException(status_code=409, detail="已解锁薪资批次缺少历史核算结果")
        after_request = _attendance_snapshot_from_body(body)
        calculation_fields = {
            "expected_days",
            "actual_days",
            "worked_hours",
            "rest_days",
            "overtime_hours",
            "holiday_worked_days",
        }
        changed_fields = {
            field for field in calculation_fields if before.get(field) != after_request.get(field)
        }
        allowed_fields = set(allowed_attendance_fields(correction_prior_result, "ATTEND_WAGE")) | {
            "overtime_hours"
        }
        unsupported_fields = changed_fields - allowed_fields
        if unsupported_fields:
            field_names = ", ".join(sorted(unsupported_fields))
            raise HTTPException(
                status_code=422,
                detail=(
                    "The historical calculation path does not use these attendance fields: "
                    f"{field_names}"
                ),
            )
    if correction_round and _attendance_calculation_values(
        before
    ) == _attendance_calculation_values(_attendance_snapshot_from_body(body)):
        raise HTTPException(status_code=422, detail="更正必须至少改变一项参与计薪的考勤数据")
    rec = _upsert_attendance(
        session,
        employee_id,
        period,
        body,
        generated_expected_days=expected_resolution.generated_expected_days,
        expected_days_rule_id=expected_resolution.expected_days_rule_id,
    )
    after = _attendance_snapshot(rec)
    if correction_round:
        if correction_batch is None:
            raise HTTPException(status_code=409, detail="已解锁薪资批次不包含该员工")
        _record_reopened_source_correction(
            session,
            batch=correction_batch,
            employee_id=employee_id,
            item="ATTENDANCE_SOURCE",
            before=before,
            after=after,
            reason=body.correction_reason or "",
            attachment_url=body.attachment_url,
            principal=principal,
        )
    if expected_days_adjusted:
        audit.record(
            session,
            action="attendance.expected_days.adjust",
            actor=(principal.user_id, principal.username),
            target_type="employee",
            target_id=employee_id,
            detail={
                "period": period,
                "before": before,
                "after": after,
                "reason": expected_resolution.audit_reason,
                "correction": correction_round,
            },
        )
    audit.record(
        session,
        action="attendance.set",
        actor=(principal.user_id, principal.username),
        target_type="employee",
        target_id=employee_id,
        detail={
            "period": period,
            "correction": correction_round,
            "before": before,
            "after": after,
            "reason": (
                body.correction_reason
                if correction_round
                else expected_resolution.audit_reason if expected_days_adjusted else None
            ),
        },
    )
    session.commit()
    return rec


@router.get("/attendance", response_model=list[AttendanceOut])
def list_attendance(
    period: str = Query(..., pattern=_PERIOD),
    principal: Principal = Depends(require_permission(Perm.ATTENDANCE_READ)),
    session: Session = Depends(get_session),
) -> list[AttendanceRecord]:
    scope = resolve_permission_org_scope(session, principal, Perm.ATTENDANCE_READ)
    stmt = (
        select(AttendanceRecord)
        .join(Employee, Employee.id == AttendanceRecord.employee_id)
        .where(AttendanceRecord.period == period, Employee.is_deleted.is_(False))
    )
    if scope is not None:
        stmt = stmt.where(Employee.org_unit_id.in_(scope))
    return list(session.scalars(stmt).all())


# ------------------- 绩效 -------------------
class PerformanceBody(BaseModel):
    coefficient: Decimal = Field(
        default=Decimal("1.000"), ge=0, le=5, max_digits=5, decimal_places=3
    )
    score: Decimal | None = Field(default=None, ge=0, le=100, max_digits=6, decimal_places=2)
    remark: str | None = Field(default=None, max_length=255)

    @field_validator("remark", mode="before")
    @classmethod
    def strip_remark(cls, value: object) -> object:
        if isinstance(value, str):
            return value.strip() or None
        return value


class PerformanceOut(BaseModel):
    employee_id: int
    period: str
    coefficient: Decimal
    score: Decimal | None
    remark: str | None

    model_config = {"from_attributes": True}


def _performance_snapshot(record: PerformanceRecord | None) -> dict[str, object]:
    if record is None:
        return {"record_exists": False}
    return {
        "record_exists": True,
        "coefficient": decimal_text(record.coefficient),
        "score": decimal_text(record.score),
        "remark": record.remark,
    }


def _ensure_performance_period_mutable(session: Session, principal: Principal, period: str) -> None:
    """Protect performance input after a batch starts or is reopened.

    A reopened attendance correction has an auditable reconciliation path.  The
    dedicated performance correction workflow has not been introduced yet, so
    neither the direct form nor a bulk file may alter a reopened batch.
    """
    if _ensure_period_mutable(session, period):
        raise HTTPException(
            status_code=409,
            detail="已重开批次的绩效数据需通过逐条、受审计的更正流程处理",
        )


@router.put("/employees/{employee_id}/performance/{period}", response_model=PerformanceOut)
def set_performance(
    employee_id: int,
    period: str,
    body: PerformanceBody,
    principal: Principal = Depends(require_permission(Perm.ATTENDANCE_WRITE)),
    session: Session = Depends(get_session),
) -> PerformanceRecord:
    if not _period_ok(period):
        raise HTTPException(status_code=422, detail="周期格式应为 YYYY-MM")
    _ensure_performance_period_mutable(session, principal, period)
    _visible_employee(
        session,
        resolve_permission_org_scope(session, principal, Perm.ATTENDANCE_WRITE),
        employee_id,
    )
    rec = session.scalars(
        select(PerformanceRecord).where(
            PerformanceRecord.employee_id == employee_id,
            PerformanceRecord.period == period,
        )
    ).first()
    before = _performance_snapshot(rec)
    values = body.model_dump()
    if rec is None:
        rec = PerformanceRecord(employee_id=employee_id, period=period, **values)
        session.add(rec)
    else:
        for field, value in values.items():
            setattr(rec, field, value)
    session.flush()
    after = _performance_snapshot(rec)
    audit.record(
        session,
        action="performance.set",
        actor=(principal.user_id, principal.username),
        target_type="employee",
        target_id=employee_id,
        detail={"period": period, "before": before, "after": after},
    )
    session.commit()
    return rec


@router.get("/performance", response_model=list[PerformanceOut])
def list_performance(
    period: str = Query(..., pattern=_PERIOD),
    principal: Principal = Depends(require_permission(Perm.ATTENDANCE_READ)),
    session: Session = Depends(get_session),
) -> list[PerformanceRecord]:
    """List performance inputs for one payroll period within the caller's scope."""
    if not _period_ok(period):
        raise HTTPException(status_code=422, detail="周期格式应为 YYYY-MM")
    scope = resolve_permission_org_scope(session, principal, Perm.ATTENDANCE_READ)
    stmt = (
        select(PerformanceRecord)
        .join(Employee, Employee.id == PerformanceRecord.employee_id)
        .where(PerformanceRecord.period == period, Employee.is_deleted.is_(False))
        .order_by(Employee.emp_no, PerformanceRecord.id)
    )
    if scope is not None:
        stmt = stmt.where(Employee.org_unit_id.in_(scope))
    return list(session.scalars(stmt).all())


class PerformanceImportResult(BaseModel):
    matched: int
    skipped: list[str]


@dataclass(frozen=True)
class _PerformanceImportRow:
    emp_no: str
    coefficient: Decimal | None
    score: Decimal | None
    remark: str | None


@router.post("/performance/import", response_model=PerformanceImportResult)
def import_performance(
    period: str = Query(..., pattern=_PERIOD),
    file: UploadFile = File(...),
    principal: Principal = Depends(require_permission(Perm.ATTENDANCE_WRITE)),
    session: Session = Depends(get_session),
) -> PerformanceImportResult:
    """Upsert an organization-scoped performance workbook as one transaction.

    ``工号`` is required for every imported row.  ``绩效系数`` is optional:
    missing or blank values keep an existing coefficient, otherwise a new
    record starts at ``1.000``.  Blank/missing score and remark cells likewise
    preserve existing values; use the single-record endpoint to clear a value.
    """
    if not _period_ok(period):
        raise HTTPException(status_code=422, detail="周期格式应为 YYYY-MM")
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx/.xlsm 文件")
    content = file.file.read(20 * 1024 * 1024 + 1)
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="文件超过 20MB 上限")

    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        raise HTTPException(status_code=400, detail="解析失败") from None

    try:
        import_rows = _parse_performance_import_rows(workbook)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from None
    finally:
        workbook.close()

    _ensure_performance_period_mutable(session, principal, period)
    scope = resolve_permission_org_scope(session, principal, Perm.ATTENDANCE_WRITE)
    emp_nos = {row.emp_no for row in import_rows}
    employee_stmt = select(Employee).where(
        Employee.emp_no.in_(emp_nos),
        Employee.is_deleted.is_(False),
    )
    if scope is not None:
        employee_stmt = employee_stmt.where(Employee.org_unit_id.in_(scope))
    employees_by_no = {
        employee.emp_no: employee for employee in session.scalars(employee_stmt).all()
    }
    employee_ids = {employee.id for employee in employees_by_no.values()}
    performance_by_employee = {
        record.employee_id: record
        for record in session.scalars(
            select(PerformanceRecord).where(
                PerformanceRecord.employee_id.in_(employee_ids or {-1}),
                PerformanceRecord.period == period,
            )
        ).all()
    }

    skipped: list[str] = []
    pending: list[tuple[Employee, PerformanceBody, PerformanceRecord | None]] = []
    for import_row in import_rows:
        employee = employees_by_no.get(import_row.emp_no)
        if employee is None:
            # Keep an out-of-scope employee indistinguishable from a missing one.
            skipped.append(import_row.emp_no)
            continue
        try:
            body = _performance_body_from_import_row(
                import_row,
                existing=performance_by_employee.get(employee.id),
            )
        except ValidationError as exc:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"工号 {import_row.emp_no} 的绩效数据格式无效：" f"{exc.errors()[0]['msg']}"
                ),
            ) from None
        pending.append((employee, body, performance_by_employee.get(employee.id)))

    # All rows are validated before the first ORM mutation: a bad row cannot
    # leave a prior row partially imported.
    for employee, body, existing in pending:
        before = _performance_snapshot(existing)
        if existing is None:
            record = PerformanceRecord(
                employee_id=employee.id,
                period=period,
                **body.model_dump(),
            )
            session.add(record)
        else:
            record = existing
            for field, value in body.model_dump().items():
                setattr(record, field, value)
        after = _performance_snapshot(record)
        audit.record(
            session,
            action="performance.import.row",
            actor=(principal.user_id, principal.username),
            target_type="employee",
            target_id=employee.id,
            detail={"period": period, "before": before, "after": after},
        )
    session.flush()
    audit.record(
        session,
        action="performance.import",
        actor=(principal.user_id, principal.username),
        detail={"period": period, "matched": len(pending), "skipped": len(skipped)},
    )
    session.commit()
    return PerformanceImportResult(matched=len(pending), skipped=skipped)


# ------------------- Excel 导入（按工号匹配员工，组织范围内 upsert）-------------------
class AttendanceImportResult(BaseModel):
    matched: int
    skipped: list[str]  # 未匹配/越权的工号


@dataclass(frozen=True)
class _AttendanceImportRow:
    emp_no: str
    expected_days: Decimal | None
    expected_days_adjust_reason: str | None
    actual_days: Decimal | None
    worked_hours: Decimal | None
    rest_days: Decimal | None
    overtime_hours: Decimal | None
    holiday_worked_days: Decimal | None
    leave_days: Decimal | None


@router.post("/attendance/import", response_model=AttendanceImportResult)
def import_attendance(
    period: str = Query(..., pattern=_PERIOD),
    file: UploadFile = File(...),
    principal: Principal = Depends(require_permission(Perm.ATTENDANCE_WRITE)),
    session: Session = Depends(get_session),
) -> AttendanceImportResult:
    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx/.xlsm 文件")
    # This endpoint intentionally stays synchronous: openpyxl and the
    # application SQLAlchemy session are synchronous APIs.  FastAPI runs a
    # normal ``def`` route in its worker threadpool, keeping a 20 MB workbook
    # parse from blocking the async event loop.
    content = file.file.read(20 * 1024 * 1024 + 1)
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="文件超过 20MB 上限")

    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        raise HTTPException(status_code=400, detail="解析失败") from None

    try:
        import_rows = _parse_attendance_import_rows(wb)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from None
    finally:
        # XLSX XML iteration is lazy.  Parse the user-controlled file before
        # taking the payroll advisory lock so a 20 MB upload cannot block every
        # source-data write while it is decoded.
        wb.close()

    if _ensure_period_mutable(session, period):
        raise HTTPException(
            status_code=409,
            detail="已解锁批次必须通过逐条更正接口修改，以保留完整审计记录",
        )

    emp_nos = {row.emp_no for row in import_rows}
    scope = resolve_permission_org_scope(session, principal, Perm.ATTENDANCE_WRITE)
    employee_stmt = select(Employee).where(
        Employee.emp_no.in_(emp_nos),
        Employee.is_deleted.is_(False),
    )
    if scope is not None:
        employee_stmt = employee_stmt.where(Employee.org_unit_id.in_(scope))
    employees_by_no = {
        employee.emp_no: employee for employee in session.scalars(employee_stmt).all()
    }
    employee_ids = {employee.id for employee in employees_by_no.values()}
    attendance_by_employee = {
        record.employee_id: record
        for record in session.scalars(
            select(AttendanceRecord).where(
                AttendanceRecord.employee_id.in_(employee_ids or {-1}),
                AttendanceRecord.period == period,
            )
        ).all()
    }

    skipped: list[str] = []
    pending: list[
        tuple[Employee, AttendanceBody, AttendanceRecord | None, _ExpectedDaysResolution]
    ] = []
    for import_row in import_rows:
        employee = employees_by_no.get(import_row.emp_no)
        if employee is None:
            # Missing and out-of-scope employees are deliberately
            # indistinguishable to preserve the organization boundary.
            skipped.append(import_row.emp_no)
            continue
        existing = attendance_by_employee.get(employee.id)
        try:
            body = _attendance_body_from_import_row(
                import_row,
                existing=existing,
                employee=employee,
            )
        except (ValidationError, ValueError) as exc:
            message = exc.errors()[0]["msg"] if isinstance(exc, ValidationError) else str(exc)
            raise HTTPException(
                status_code=422,
                detail=f"工号 {import_row.emp_no} 的考勤数据格式无效：{message}",
            ) from None
        expected_resolution = _resolve_expected_days_for_write(
            session,
            employee=employee,
            period=period,
            body=body,
            existing=existing,
            principal=principal,
        )
        body = expected_resolution.body
        pending.append((employee, body, existing, expected_resolution))

    for employee, body, existing, expected_resolution in pending:
        before = _attendance_snapshot(existing)
        expected_days_adjusted = expected_resolution.audit_reason is not None
        record = _upsert_attendance(
            session,
            employee.id,
            period,
            body,
            generated_expected_days=expected_resolution.generated_expected_days,
            expected_days_rule_id=expected_resolution.expected_days_rule_id,
            existing=existing,
            already_loaded=True,
            flush=False,
        )
        after = _attendance_snapshot(record)
        audit.record(
            session,
            action="attendance.import.row",
            actor=(principal.user_id, principal.username),
            target_type="employee",
            target_id=employee.id,
            detail={"period": period, "before": before, "after": after},
        )
        if expected_days_adjusted:
            audit.record(
                session,
                action="attendance.expected_days.adjust",
                actor=(principal.user_id, principal.username),
                target_type="employee",
                target_id=employee.id,
                detail={
                    "period": period,
                    "before": before,
                    "after": after,
                    "reason": expected_resolution.audit_reason,
                    "correction": False,
                    "source": "attendance_import",
                },
            )
    session.flush()
    audit.record(
        session,
        action="attendance.import",
        actor=(principal.user_id, principal.username),
        detail={"period": period, "matched": len(pending), "skipped": len(skipped)},
    )
    session.commit()
    return AttendanceImportResult(matched=len(pending), skipped=skipped)


def _period_ok(period: str) -> bool:
    import re

    return bool(re.fullmatch(_PERIOD, period)) and 1 <= int(period[-2:]) <= 12


def _parse_attendance_import_rows(workbook) -> list[_AttendanceImportRow]:
    rows: list[_AttendanceImportRow] = []
    seen_emp_nos: set[str] = set()
    for sheet in workbook.worksheets:
        headers = [clean_text(cell.value) for cell in sheet[1]]
        columns = {header: index for index, header in enumerate(headers)}
        if "工号" not in columns:
            continue
        for cells in sheet.iter_rows(min_row=2):
            emp_no = clean_text(cells[columns["工号"]].value)
            if not emp_no:
                continue
            if emp_no in seen_emp_nos:
                raise ValueError(f"导入文件中工号 {emp_no} 重复")
            seen_emp_nos.add(emp_no)
            rows.append(
                _AttendanceImportRow(
                    emp_no=emp_no,
                    expected_days=_attendance_import_decimal(
                        cells,
                        columns,
                        emp_no=emp_no,
                        label="应出勤",
                        names=("应出勤",),
                    ),
                    expected_days_adjust_reason=_import_text(cells, columns, "应出勤调整原因"),
                    actual_days=_attendance_import_decimal(
                        cells,
                        columns,
                        emp_no=emp_no,
                        label="实出勤",
                        names=("实出勤",),
                    ),
                    worked_hours=_attendance_import_decimal(
                        cells,
                        columns,
                        emp_no=emp_no,
                        label="出勤工时",
                        names=("出勤工时", "工时"),
                    ),
                    rest_days=_attendance_import_decimal(
                        cells,
                        columns,
                        emp_no=emp_no,
                        label="休息天数",
                        names=("休息天数", "休息"),
                    ),
                    overtime_hours=_attendance_import_decimal(
                        cells,
                        columns,
                        emp_no=emp_no,
                        label="加班",
                        names=("加班",),
                    ),
                    holiday_worked_days=_attendance_import_decimal(
                        cells,
                        columns,
                        emp_no=emp_no,
                        label="法定节假日出勤天数",
                        names=("法定节假日出勤天数", "法定出勤"),
                    ),
                    leave_days=_attendance_import_decimal(
                        cells,
                        columns,
                        emp_no=emp_no,
                        label="请假",
                        names=("请假",),
                    ),
                )
            )
    return rows


def _attendance_import_decimal(
    cells,
    columns: dict[str, int],
    *,
    emp_no: str,
    label: str,
    names: tuple[str, ...],
) -> Decimal | None:
    """Parse an optional number without treating malformed input as blank."""
    for name in names:
        index = columns.get(name)
        if index is None or index >= len(cells):
            continue
        raw_value = cells[index].value
        if raw_value is None or not clean_text(raw_value):
            continue
        value = parse_money(raw_value)
        if value is None:
            raise ValueError(f"工号 {emp_no} 的 {label} 格式无效")
        return value
    return None


def _import_text(cells, columns: dict[str, int], *names: str) -> str | None:
    for name in names:
        index = columns.get(name)
        if index is None or index >= len(cells):
            continue
        value = clean_text(cells[index].value)
        if value:
            return value
    return None


def _parse_performance_import_rows(workbook) -> list[_PerformanceImportRow]:
    """Parse every performance sheet before any database write.

    A sheet must contain ``工号`` and at least one supported performance field
    to participate.  This lets a workbook include a cover/reference sheet
    without accidentally creating default records from its employee list.
    """
    rows: list[_PerformanceImportRow] = []
    seen_emp_nos: set[str] = set()
    found_performance_sheet = False
    supported_fields = (
        "绩效系数",
        "系数",
        "绩效得分",
        "得分",
        "备注",
        "绩效备注",
    )
    for sheet in workbook.worksheets:
        headers = [clean_text(cell.value) for cell in sheet[1]]
        columns = {header: index for index, header in enumerate(headers)}
        if "工号" not in columns or not any(name in columns for name in supported_fields):
            continue
        found_performance_sheet = True
        for cells in sheet.iter_rows(min_row=2):
            emp_no = clean_text(cells[columns["工号"]].value)
            if not emp_no:
                continue
            if emp_no in seen_emp_nos:
                raise ValueError(f"导入文件中工号 {emp_no} 重复")
            seen_emp_nos.add(emp_no)
            rows.append(
                _PerformanceImportRow(
                    emp_no=emp_no,
                    coefficient=_performance_import_decimal(
                        cells,
                        columns,
                        emp_no=emp_no,
                        label="绩效系数",
                        names=("绩效系数", "系数"),
                    ),
                    score=_performance_import_decimal(
                        cells,
                        columns,
                        emp_no=emp_no,
                        label="绩效得分",
                        names=("绩效得分", "得分"),
                    ),
                    remark=_import_text(cells, columns, "备注", "绩效备注"),
                )
            )
    if not found_performance_sheet:
        raise ValueError("导入文件需包含工号及至少一项绩效系数、绩效得分或备注字段")
    return rows


def _performance_import_decimal(
    cells,
    columns: dict[str, int],
    *,
    emp_no: str,
    label: str,
    names: tuple[str, ...],
) -> Decimal | None:
    """Parse an optional numeric cell without turning malformed input into blank."""
    for name in names:
        index = columns.get(name)
        if index is None or index >= len(cells):
            continue
        raw_value = cells[index].value
        if raw_value is None or not clean_text(raw_value):
            continue
        value = parse_money(raw_value)
        if value is None:
            raise ValueError(f"工号 {emp_no} 的 {label} 格式无效")
        return value
    return None


def _performance_body_from_import_row(
    import_row: _PerformanceImportRow,
    *,
    existing: PerformanceRecord | None,
) -> PerformanceBody:
    """Apply only values supplied by the file and preserve existing optional data."""
    return PerformanceBody(
        coefficient=(
            import_row.coefficient
            if import_row.coefficient is not None
            else existing.coefficient if existing is not None else Decimal("1.000")
        ),
        score=(
            import_row.score
            if import_row.score is not None
            else existing.score if existing is not None else None
        ),
        remark=(
            import_row.remark
            if import_row.remark is not None
            else existing.remark if existing is not None else None
        ),
    )


def _required_import_decimal(value: Decimal | None, employee: Employee, label: str) -> Decimal:
    if value is None:
        raise ValueError(f"工号 {employee.emp_no} 缺少 {label}")
    return value


@overload
def _preserved_import_decimal(
    value: Decimal | None,
    existing: AttendanceRecord | None,
    field: Literal["worked_hours"],
) -> Decimal | None: ...


@overload
def _preserved_import_decimal(
    value: Decimal | None,
    existing: AttendanceRecord | None,
    field: Literal["rest_days", "overtime_hours", "holiday_worked_days", "leave_days"],
) -> Decimal: ...


def _preserved_import_decimal(
    value: Decimal | None,
    existing: AttendanceRecord | None,
    field: str,
) -> Decimal | None:
    if value is not None:
        return value
    if existing is not None:
        return getattr(existing, field)
    return None if field == "worked_hours" else Decimal(0)


def _attendance_body_from_import_row(
    import_row: _AttendanceImportRow,
    *,
    existing: AttendanceRecord | None,
    employee: Employee,
) -> AttendanceBody:
    """Build a validated row without erasing calculation inputs not in a file.

    New hourly employees must supply their worked hours.  Existing records may
    omit optional columns because the import then preserves the persisted value.
    """
    worked_hours = _preserved_import_decimal(import_row.worked_hours, existing, "worked_hours")
    uses_approved_days = employee.is_special_position or requires_approved_attendance_days(
        employee.position_title
    )
    if (
        existing is None
        and not uses_approved_days
        and employee.department.value in {"DINING", "KITCHEN"}
        and import_row.worked_hours is None
    ):
        raise ValueError(f"工号 {employee.emp_no} 的工时制岗位缺少出勤工时")
    return AttendanceBody(
        expected_days=_required_import_decimal(import_row.expected_days, employee, "应出勤"),
        expected_days_adjust_reason=import_row.expected_days_adjust_reason,
        actual_days=_required_import_decimal(import_row.actual_days, employee, "实出勤"),
        worked_hours=worked_hours,
        rest_days=_preserved_import_decimal(import_row.rest_days, existing, "rest_days"),
        overtime_hours=_preserved_import_decimal(
            import_row.overtime_hours, existing, "overtime_hours"
        ),
        holiday_worked_days=_preserved_import_decimal(
            import_row.holiday_worked_days, existing, "holiday_worked_days"
        ),
        leave_days=_preserved_import_decimal(import_row.leave_days, existing, "leave_days"),
        late_count=existing.late_count if existing is not None else 0,
        early_leave_count=existing.early_leave_count if existing is not None else 0,
    )
