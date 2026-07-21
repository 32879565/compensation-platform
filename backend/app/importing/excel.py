"""Excel 工资表读取（openpyxl）→ SalaryRow 列表。

聚焦标准结构：某行含"姓名"表头，其下为数据行；门店名取工作表首行或 sheet 名。
表头经 normalize_header 归一化，金额经 parse_money（失败置 None，由上层报错）。
"""

from __future__ import annotations

import re
import struct
from collections.abc import Mapping, Sequence
from contextlib import contextmanager
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import IO
from zipfile import ZipFile

from openpyxl import load_workbook

from app.importing.header_rules import (
    MONEY_FIELDS,
    SKIP_SHEET_TITLES,
    SUM_MERGED_FIELD_LABELS,
    SUMMARY_ROW_NAMES,
)
from app.importing.parser import (
    SalaryRow,
    auto_store_aliases,
    clean_text,
    normalize_emp_no,
    normalize_header,
    normalize_store_name,
    parse_money,
    standard_store_name,
)
from app.importing.store_aliases import STORE_ALIASES

_SUMMARY_KEYWORDS = ("汇总", "总表", "合计", "初版", "一厅面")
_INVALID_STORE_TITLES = {"", "序号", "门店", "Sheet1", "`", "姓名"}
MAX_EMPLOYEE_ROWS = 10_000
MAX_XLSX_ARCHIVE_MEMBERS = 1_000
MAX_XLSX_EXPANDED_BYTES = 64 * 1024 * 1024
MAX_XLSX_MEMBER_BYTES = 64 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 200
MAX_WORKSHEET_ROWS = 20_000
MAX_WORKSHEET_COLUMNS = 256
MAX_WORKSHEET_DIMENSION_CELLS = 2_000_000
MAX_WORKBOOK_SHEETS = 128
MAX_WORKBOOK_STREAMED_CELLS = 5_000_000
MAX_CELL_TEXT_LENGTH = 8_192


class WorkbookLimitError(ValueError):
    """The workbook exceeds a bounded parsing safety limit."""


@dataclass
class ReadResult:
    rows: list[SalaryRow]
    warnings: list[str]


@contextmanager
def _archive_stream(source: str | IO[bytes]):
    """Expose a seekable stream without changing a caller-owned stream position."""
    if isinstance(source, (str, Path)):
        with open(source, "rb") as stream:
            yield stream
        return

    position = source.tell()
    source.seek(0)
    try:
        yield source
    finally:
        source.seek(position)


def _zip_directory_metadata(source: IO[bytes]) -> tuple[int, int, int]:
    """Return declared count, offset and size from a non-ZIP64 EOCD record."""
    source.seek(0, 2)
    archive_size = source.tell()
    tail_size = min(archive_size, 65_535 + 22)
    source.seek(archive_size - tail_size)
    tail = source.read(tail_size)
    signature = b"PK\x05\x06"
    cursor = len(tail)
    while cursor:
        offset = tail.rfind(signature, 0, cursor)
        if offset < 0:
            break
        if offset + 22 <= len(tail):
            comment_length = struct.unpack_from("<H", tail, offset + 20)[0]
            if offset + 22 + comment_length == len(tail):
                disk, directory_disk, disk_entries, total_entries = struct.unpack_from(
                    "<4H", tail, offset + 4
                )
                if disk != 0 or directory_disk != 0 or disk_entries != total_entries:
                    raise WorkbookLimitError("Multi-disk XLSX archives are not supported")
                if total_entries == 0xFFFF:
                    raise WorkbookLimitError("ZIP64 XLSX archives are not supported")
                directory_size, directory_offset = struct.unpack_from("<2L", tail, offset + 12)
                eocd_offset = archive_size - tail_size + offset
                if eocd_offset >= 20:
                    source.seek(eocd_offset - 20)
                    if source.read(4) == b"PK\x06\x07":
                        raise WorkbookLimitError("ZIP64 XLSX archives are not supported")
                if directory_offset + directory_size != eocd_offset:
                    raise WorkbookLimitError("XLSX central directory boundaries are invalid")
                return int(total_entries), int(directory_offset), int(directory_size)
        cursor = offset
    raise WorkbookLimitError("XLSX ZIP directory is invalid")


def _count_central_directory_entries(
    source: IO[bytes], *, directory_offset: int, directory_size: int
) -> int:
    """Stream raw central-directory headers with a cap before constructing ZipFile."""
    source.seek(directory_offset)
    remaining = directory_size
    count = 0
    while remaining:
        if remaining < 46:
            raise WorkbookLimitError("XLSX central directory record is truncated")
        header = source.read(46)
        if len(header) != 46 or header[:4] != b"PK\x01\x02":
            raise WorkbookLimitError("XLSX central directory record is invalid")
        name_length, extra_length, comment_length = struct.unpack_from("<3H", header, 28)
        variable_length = name_length + extra_length + comment_length
        record_length = 46 + variable_length
        if record_length > remaining:
            raise WorkbookLimitError("XLSX central directory record exceeds its boundary")
        source.seek(variable_length, 1)
        remaining -= record_length
        count += 1
        if count > MAX_XLSX_ARCHIVE_MEMBERS:
            raise WorkbookLimitError(
                "XLSX archive member count exceeds " f"the limit of {MAX_XLSX_ARCHIVE_MEMBERS}"
            )
    return count


def _validate_xlsx_archive(source: str | IO[bytes]) -> None:
    """Reject ZIP containers that could expand beyond bounded parser resources."""
    with _archive_stream(source) as archive_source:
        declared_count, directory_offset, directory_size = _zip_directory_metadata(archive_source)
        if declared_count > MAX_XLSX_ARCHIVE_MEMBERS:
            raise WorkbookLimitError(
                "XLSX archive member count exceeds " f"the limit of {MAX_XLSX_ARCHIVE_MEMBERS}"
            )
        actual_count = _count_central_directory_entries(
            archive_source,
            directory_offset=directory_offset,
            directory_size=directory_size,
        )
        if actual_count != declared_count:
            raise WorkbookLimitError("XLSX central directory count does not match EOCD")
        archive_source.seek(0)
        with ZipFile(archive_source) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_ARCHIVE_MEMBERS:
                raise WorkbookLimitError(
                    "XLSX archive member count exceeds " f"the limit of {MAX_XLSX_ARCHIVE_MEMBERS}"
                )

            expanded_bytes = 0
            for member in members:
                if member.file_size > MAX_XLSX_MEMBER_BYTES:
                    raise WorkbookLimitError(
                        "XLSX archive member expanded size exceeds "
                        f"the limit of {MAX_XLSX_MEMBER_BYTES} bytes"
                    )
                expanded_bytes += member.file_size
                if expanded_bytes > MAX_XLSX_EXPANDED_BYTES:
                    raise WorkbookLimitError(
                        "XLSX archive expanded size exceeds "
                        f"the limit of {MAX_XLSX_EXPANDED_BYTES} bytes"
                    )
                if member.file_size > max(1, member.compress_size) * MAX_XLSX_COMPRESSION_RATIO:
                    raise WorkbookLimitError(
                        "XLSX archive compression ratio exceeds "
                        f"the limit of {MAX_XLSX_COMPRESSION_RATIO}:1"
                    )


def _validate_cell_value(value: object) -> None:
    if isinstance(value, str) and len(value) > MAX_CELL_TEXT_LENGTH:
        raise WorkbookLimitError(
            f"XLSX cell text exceeds the limit of {MAX_CELL_TEXT_LENGTH} characters"
        )


def _validate_worksheet_dimensions(sheet) -> None:
    max_row = int(sheet.max_row or 0)
    max_column = int(sheet.max_column or 0)
    if max_row > MAX_WORKSHEET_ROWS:
        raise WorkbookLimitError(
            f"Worksheet row dimension exceeds the limit of {MAX_WORKSHEET_ROWS}"
        )
    if max_column > MAX_WORKSHEET_COLUMNS:
        raise WorkbookLimitError(
            f"Worksheet column dimension exceeds the limit of {MAX_WORKSHEET_COLUMNS}"
        )
    if max_row * max_column > MAX_WORKSHEET_DIMENSION_CELLS:
        raise WorkbookLimitError(
            "Worksheet cell dimension exceeds "
            f"the limit of {MAX_WORKSHEET_DIMENSION_CELLS} cells"
        )


def _validate_streamed_row(
    row_number: int,
    cells: Sequence[object],
    *,
    processed_cells: int,
) -> int:
    """Enforce actual streamed complexity even when XLSX dimension metadata lies."""
    actual_row = row_number
    actual_column = len(cells)
    for index, cell in enumerate(cells, start=1):
        cell_row = getattr(cell, "row", None)
        cell_column = getattr(cell, "column", None)
        if isinstance(cell_row, int):
            actual_row = max(actual_row, cell_row)
        if isinstance(cell_column, int):
            actual_column = max(actual_column, cell_column)
        else:
            actual_column = max(actual_column, index)
    if actual_row > MAX_WORKSHEET_ROWS:
        raise WorkbookLimitError(
            f"Worksheet row dimension exceeds the limit of {MAX_WORKSHEET_ROWS}"
        )
    if actual_column > MAX_WORKSHEET_COLUMNS:
        raise WorkbookLimitError(
            f"Worksheet column dimension exceeds the limit of {MAX_WORKSHEET_COLUMNS}"
        )
    next_processed_cells = processed_cells + actual_column
    if next_processed_cells > MAX_WORKSHEET_DIMENSION_CELLS:
        raise WorkbookLimitError(
            "Worksheet cell dimension exceeds "
            f"the limit of {MAX_WORKSHEET_DIMENSION_CELLS} cells"
        )
    return next_processed_cells


def _accumulate_workbook_cells(processed_cells: int, *, additional_cells: int) -> int:
    """Bound cumulative streamed work across every worksheet in one workbook."""
    next_processed_cells = processed_cells + additional_cells
    if next_processed_cells > MAX_WORKBOOK_STREAMED_CELLS:
        raise WorkbookLimitError(
            "XLSX workbook cell budget exceeds " f"the limit of {MAX_WORKBOOK_STREAMED_CELLS} cells"
        )
    return next_processed_cells


def _find_header_row(sheet, max_scan: int = 5) -> int | None:
    for r in range(1, max_scan + 1):
        cells = sheet[r]
        _validate_streamed_row(r, cells, processed_cells=0)
        for cell in cells:
            _validate_cell_value(cell.value)
        values = [clean_text(c.value) for c in cells]
        if "姓名" in values:
            return r
    return None


def _sheet_store_name(sheet) -> str:
    cells = sheet[1]
    _validate_streamed_row(1, cells, processed_cells=0)
    for cell in cells:
        _validate_cell_value(cell.value)
    first = [clean_text(c.value) for c in cells if clean_text(c.value)]
    if first and first[0] not in _INVALID_STORE_TITLES:
        return normalize_store_name(first[0])
    return normalize_store_name(sheet.title)


def _has_month_salary(headers: list[str], month: int | None) -> bool:
    if month is None:
        return False
    pat = re.compile(rf"{month}月\s*(综合薪资|薪资|工资|底薪)")
    return any(pat.fullmatch(re.sub(r"\s+", "", h)) for h in headers)


def _merge_cell(
    fields: dict[str, str], money: dict[str, object], canonical: str, raw: object
) -> None:
    """把一列的值合入记录，正确处理同名列碰撞（移植旧 add_record_value 语义）。

    - 金额：空/无法解析(None) 绝不覆盖已有真实值；SUM 字段跨列累加，其余 last-non-empty。
    - 文本：last-non-empty。
    """
    text = clean_text(raw)
    if text:
        fields[canonical] = text
    elif canonical not in fields:
        fields[canonical] = text

    if canonical not in MONEY_FIELDS:
        return
    value = parse_money(raw)
    if value is None:
        return  # 空/不可解析的后列不得把前列真实金额覆盖成 None
    existing = money.get(canonical)
    if canonical in SUM_MERGED_FIELD_LABELS and isinstance(existing, Decimal):
        money[canonical] = existing + value
    else:
        money[canonical] = value


def read_salary_workbook(
    source: str | IO[bytes],
    *,
    period: str,
    aliases: Mapping[str, str] | None = None,
) -> ReadResult:
    # The legacy mappings are part of the import contract.  Explicit mappings are an overlay
    # so a deployment can add a newly approved alias without losing the baseline catalogue.
    aliases = {**STORE_ALIASES, **(aliases or {})}
    month = int(period.split("-")[1]) if "-" in period else None
    _validate_xlsx_archive(source)
    wb = load_workbook(source, read_only=True, data_only=True)
    rows: list[SalaryRow] = []
    warnings: list[str] = []

    try:
        if len(wb.worksheets) > MAX_WORKBOOK_SHEETS:
            raise WorkbookLimitError(
                f"XLSX worksheet count exceeds the limit of {MAX_WORKBOOK_SHEETS}"
            )
        workbook_processed_cells = 0
        for sheet in wb.worksheets:
            title = sheet.title
            if any(k in title for k in _SUMMARY_KEYWORDS) or title in SKIP_SHEET_TITLES:
                continue
            _validate_worksheet_dimensions(sheet)
            # Read-only openpyxl trusts the optional XML <dimension> when
            # deciding where iteration ends.  Reset it after rejecting an
            # over-reported dimension so under-reporting cannot hide rows from
            # the actual streamed row/cell limits below.
            sheet.reset_dimensions()
            header_row = _find_header_row(sheet)
            if header_row is None:
                warnings.append(f"工作表『{title}』未找到姓名列，已跳过")
                continue

            headers = [clean_text(c.value) for c in sheet[header_row]]
            has_month = _has_month_salary(headers, month)
            col_map: dict[int, str] = {}
            name_col: int | None = None
            emp_no_col: int | None = None
            for idx, raw in enumerate(headers, start=1):
                if raw == "姓名":
                    name_col = idx
                    continue
                if raw in ("工号", "员工编号", "员工号"):
                    emp_no_col = idx
                    continue
                canonical = normalize_header(
                    raw, month=month, column_index=idx, has_month_salary=has_month
                )
                if canonical:
                    col_map[idx] = canonical
            if name_col is None:
                continue

            store_name = standard_store_name(_sheet_store_name(sheet), aliases)

            processed_cells = 0
            for row_number, excel_row in enumerate(
                sheet.iter_rows(min_row=header_row + 1), start=header_row + 1
            ):
                previous_processed_cells = processed_cells
                processed_cells = _validate_streamed_row(
                    row_number,
                    excel_row,
                    processed_cells=processed_cells,
                )
                workbook_processed_cells = _accumulate_workbook_cells(
                    workbook_processed_cells,
                    additional_cells=processed_cells - previous_processed_cells,
                )
                for cell in excel_row:
                    _validate_cell_value(cell.value)
                name = clean_text(excel_row[name_col - 1].value) if name_col else ""
                if not name or name == "姓名" or name in SUMMARY_ROW_NAMES:
                    continue
                if len(rows) >= MAX_EMPLOYEE_ROWS:
                    raise WorkbookLimitError(
                        f"Workbook employee row count exceeds the limit of {MAX_EMPLOYEE_ROWS}"
                    )
                emp_no = normalize_emp_no(excel_row[emp_no_col - 1].value) if emp_no_col else None
                fields: dict[str, str] = {}
                money: dict[str, object] = {}
                for idx, canonical in col_map.items():
                    if idx - 1 >= len(excel_row):
                        continue
                    _merge_cell(fields, money, canonical, excel_row[idx - 1].value)
                rows.append(
                    SalaryRow(
                        period=period,
                        name=name,
                        store_name=store_name,
                        emp_no=emp_no,
                        fields=fields,
                        money=money,  # type: ignore[arg-type]
                    )
                )
        # Preserve the old parser's low-risk automatic ``X`` → ``X店`` repair as well.  It runs
        # after all sheets are available, so an alias can be inferred from a canonical sibling.
        automatic_aliases = auto_store_aliases({row.store_name for row in rows if row.store_name})
        for row in rows:
            row.store_name = standard_store_name(
                automatic_aliases.get(row.store_name, row.store_name), aliases
            )
        return ReadResult(rows=rows, warnings=warnings)
    finally:
        wb.close()
